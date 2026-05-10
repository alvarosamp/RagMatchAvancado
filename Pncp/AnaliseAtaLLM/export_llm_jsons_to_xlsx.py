"""Pncp/AnaliseAtaLLM/export_llm_jsons_to_xlsx.py

Exporta resultados já gerados (JSON) para uma única planilha .xlsx,
com uma aba por arquivo JSON e 1 linha por item.

Default: pega 10 arquivos em Pncp/AnaliseAtaLLM/results_llm/*_llm.json

Como rodar (recomendado usando o venv):
  .\.venv\Scripts\python.exe Pncp/AnaliseAtaLLM/export_llm_jsons_to_xlsx.py

Saída padrão:
  Pncp/AnaliseAtaLLM/resultados/xlsx/atas_llm_10_from_json.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook


_INVALID_SHEET_CHARS = r"[\[\]:*?/\\]"


def _sanitize_sheet_name(name: str) -> str:
    name = re.sub(_INVALID_SHEET_CHARS, "_", name or "aba")
    name = re.sub(r"\s+", " ", name).strip()
    if not name:
        name = "aba"
    return name[:31]


def _safe_cell(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value)
    if len(s) > 32767:
        s = s[:32760] + "..."
    return s


def _read_json(path: Path) -> dict:
    # Alguns arquivos podem ter BOM.
    text = path.read_text(encoding="utf-8-sig")
    data = json.loads(text)
    if isinstance(data, dict):
        return data
    # fallback mínimo
    return {"itens": data}


def main() -> int:
    parser = argparse.ArgumentParser(description="Exporta 10 JSONs LLM para XLSX (abas por arquivo).")
    parser.add_argument(
        "--input-glob",
        default=str(Path(__file__).resolve().parent / "results_llm" / "*_llm.json"),
        help="Glob dos JSONs de entrada (default: results_llm/*_llm.json)",
    )
    parser.add_argument("--limit", type=int, default=10, help="Quantos JSONs exportar (default: 10)")
    parser.add_argument(
        "--output",
        default=str(Path(__file__).resolve().parent / "resultados" / "xlsx" / "atas_llm_10_from_json.xlsx"),
        help="Caminho do XLSX de saída",
    )
    args = parser.parse_args()

    paths = sorted(Path().glob(args.input_glob))[: max(0, args.limit)]
    if not paths:
        raise SystemExit(f"Nenhum arquivo encontrado para o glob: {args.input_glob}")

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    columns = [
        "arquivo_json",
        "id_pncp",
        "numero_ata",
        "orgao",
        "data_assinatura",
        "vigencia",
        "objeto",
        "tokens_usados",
        "aviso",
        "numero_item",
        "descricao",
        "tipo",
        "marca",
        "modelo",
        "quantidade",
        "unidade",
        "valor_unitario",
        "valor_total",
        "fornecedor",
        "cnpj_fornecedor",
        "especificacoes",
        "observacoes",
        "raw_descricao",
    ]

    for i, p in enumerate(paths, 1):
        data = _read_json(p)

        sheet_name = _sanitize_sheet_name(p.stem)
        base_name = sheet_name
        suffix = 2
        while sheet_name in wb.sheetnames:
            sheet_name = _sanitize_sheet_name(f"{base_name}_{suffix}")
            suffix += 1

        ws = wb.create_sheet(title=sheet_name)
        ws.append(columns)
        ws.freeze_panes = "A2"

        itens = data.get("itens") or []
        if not isinstance(itens, list):
            itens = []

        for it in itens:
            it = it or {}
            row = {
                "arquivo_json": p.name,
                "id_pncp": data.get("id_pncp"),
                "numero_ata": data.get("numero_ata"),
                "orgao": data.get("orgao"),
                "data_assinatura": data.get("data_assinatura"),
                "vigencia": data.get("vigencia"),
                "objeto": data.get("objeto"),
                "tokens_usados": data.get("tokens_usados"),
                "aviso": data.get("aviso"),
                "numero_item": it.get("numero_item"),
                "descricao": it.get("descricao"),
                "tipo": it.get("tipo"),
                "marca": it.get("marca"),
                "modelo": it.get("modelo"),
                "quantidade": it.get("quantidade"),
                "unidade": it.get("unidade"),
                "valor_unitario": it.get("valor_unitario"),
                "valor_total": it.get("valor_total"),
                "fornecedor": it.get("fornecedor"),
                "cnpj_fornecedor": it.get("cnpj_fornecedor"),
                "especificacoes": "; ".join((it.get("especificacoes") or []) if isinstance(it.get("especificacoes"), list) else []),
                "observacoes": it.get("observacoes"),
                "raw_descricao": it.get("raw_descricao"),
            }
            ws.append([_safe_cell(row[c]) for c in columns])

        print(f"[{i}/{len(paths)}] {p.name} -> aba '{sheet_name}' | itens: {len(itens)}")

    wb.save(out_path)
    print(f"OK: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
