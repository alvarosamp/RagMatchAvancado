"""Pncp/AnaliseAtaLLM/run_10_atas_to_xlsx.py

Gera uma única planilha .xlsx com 10 abas (uma por arquivo),
com 1 linha por item extraído do pipeline LLM.

Entrada padrão:
  Pncp/AnaliseAtaLLM/textos_md/*.md

Saída padrão:
  Pncp/AnaliseAtaLLM/resultados/xlsx/atas_llm_10.xlsx

Como rodar:
  python Pncp/AnaliseAtaLLM/run_10_atas_to_xlsx.py

Requisitos:
- Ollama rodando em http://localhost:11434 (ou OLLAMA_HOST)
- Modelo disponível (OLLAMA_MODEL)
"""

from __future__ import annotations

import argparse
import re
import time
from dataclasses import asdict
from pathlib import Path

from openpyxl import Workbook

from pipelinellm_prompt_ajustado import analisar_ata


_INVALID_SHEET_CHARS = r"[\[\]:*?/\\]"


def _sanitize_sheet_name(name: str) -> str:
    name = re.sub(_INVALID_SHEET_CHARS, "_", name or "aba")
    name = re.sub(r"\s+", " ", name).strip()
    if not name:
        name = "aba"
    return name[:31]


def _safe_cell(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value)
    # Limite do Excel por célula (caracteres). Evita exception no openpyxl.
    if len(s) > 32767:
        s = s[:32760] + "..."
    return s


def _write_header(ws, columns: list[str]) -> None:
    ws.append(columns)
    ws.freeze_panes = "A2"


def _write_item_rows(ws, base: dict, itens: list[dict], source_file: str) -> int:
    count = 0
    for it in itens:
        row = {
            "arquivo": source_file,
            "id_pncp": base.get("id_pncp"),
            "numero_ata": base.get("numero_ata"),
            "orgao": base.get("orgao"),
            "data_assinatura": base.get("data_assinatura"),
            "vigencia": base.get("vigencia"),
            "objeto": base.get("objeto"),
            "tokens_usados": base.get("tokens_usados"),
            "aviso": base.get("aviso"),
            "numero_item": it.get("numero_item"),
            "descricao": it.get("descricao"),
            "tipo": it.get("tipo"),
            "marca": it.get("marca"),
            "modelo": it.get("modelo"),
            "quantidade": it.get("quantidade"),
            "unidade": it.get("unidade"),
            "valor_unitario": it.get("valor_unitario"),
            "valor_total": it.get("valor_total"),
            "fornecedor": it.get("fornecedor"),
            "cnpj_fornecedor": it.get("cnpj_fornecedor"),
            "especificacoes": "; ".join(it.get("especificacoes") or []),
            "observacoes": it.get("observacoes"),
            "raw_descricao": it.get("raw_descricao"),
        }

        ws.append([_safe_cell(row[c]) for c in ws._columns_order])
        count += 1

    return count


def main() -> int:
    parser = argparse.ArgumentParser(description="Roda 10 atas e exporta para XLSX (abas por arquivo).")
    parser.add_argument(
        "--input-dir",
        default=str(Path(__file__).resolve().parent / "textos_md"),
        help="Diretório com .md (default: Pncp/AnaliseAtaLLM/textos_md)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=10,
        help="Quantos arquivos processar (default: 10)",
    )
    parser.add_argument(
        "--output",
        default=str(Path(__file__).resolve().parent / "resultados" / "xlsx" / "atas_llm_10.xlsx"),
        help="Caminho do .xlsx de saída",
    )
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    arquivos = sorted(input_dir.glob("*.md"))[: max(0, args.limit)]
    if not arquivos:
        raise SystemExit(f"Nenhum .md encontrado em {input_dir}")

    wb = Workbook()
    # Remove a aba padrão
    wb.remove(wb.active)

    columns = [
        "arquivo",
        "id_pncp",
        "numero_ata",
        "orgao",
        "data_assinatura",
        "vigencia",
        "objeto",
        "tokens_usados",
        "aviso",
        "numero_item",
        "descricao",
        "tipo",
        "marca",
        "modelo",
        "quantidade",
        "unidade",
        "valor_unitario",
        "valor_total",
        "fornecedor",
        "cnpj_fornecedor",
        "especificacoes",
        "observacoes",
        "raw_descricao",
    ]

    # hack leve: guardar a ordem das colunas no worksheet
    def set_columns_order(ws):
        ws._columns_order = columns  # type: ignore[attr-defined]

    t0 = time.time()
    for i, path in enumerate(arquivos, 1):
        stem = path.stem
        sheet_name = _sanitize_sheet_name(stem)

        # Garantir unicidade do nome da aba
        base_name = sheet_name
        suffix = 2
        while sheet_name in wb.sheetnames:
            sheet_name = _sanitize_sheet_name(f"{base_name}_{suffix}")
            suffix += 1

        ws = wb.create_sheet(title=sheet_name)
        set_columns_order(ws)
        _write_header(ws, columns)

        texto = path.read_text(encoding="utf-8")
        print(f"[{i}/{len(arquivos)}] Processando: {path.name} -> aba: {sheet_name}")

        try:
            resultado = analisar_ata(texto, id_pncp=stem, persistir=False)
            base = asdict(resultado)
            itens = base.get("itens") or []
            n = _write_item_rows(ws, base, itens, source_file=path.name)
            print(f"  itens: {n} | tokens: {base.get('tokens_usados')}")
        except Exception as e:
            # Se falhar, registra a falha como 1 linha (para não perder a aba)
            ws.append([
                path.name,
                stem,
                None,
                None,
                None,
                None,
                None,
                None,
                f"erro: {e}",
            ] + [None] * (len(columns) - 9))
            print(f"  ERRO: {e}")

    wb.save(out_path)
    elapsed = time.time() - t0
    print(f"OK: {out_path} | {len(arquivos)} abas | {elapsed:.1f}s")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
