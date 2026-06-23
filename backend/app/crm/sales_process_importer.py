from __future__ import annotations

import hashlib
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, TYPE_CHECKING

from openpyxl import load_workbook

if TYPE_CHECKING:
    from app.auth.models import Tenant, User
    from app.crm.models import CrmNotice


PRIMARY_EMAIL = "alvaroscareli@gmail.com"
FALLBACK_EMAIL = "vish88@outlook.com"
DEFAULT_PASSWORD = "Elainema157!"
DEFAULT_TENANT_SLUG = "tor-tec"
DEFAULT_TENANT_NAME = "Tor Tec"
DEFAULT_USER_NAME = "Tor Tec"


@dataclass
class ImportContext:
    tenant: "Tenant"
    user: "User"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def optional_text(value: Any) -> str | None:
    text = normalize_text(value)
    return text or None


def optional_meaningful_text(value: Any) -> str | None:
    return normalize_text(value) if has_meaningful_value(value) else None


def normalize_key_part(value: Any) -> str:
    text = normalize_text(value)
    return " ".join(text.lower().split())


def canonical_header(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    ascii_text = (
        unicodedata.normalize("NFKD", text)
        .encode("ascii", "ignore")
        .decode("ascii")
        .lower()
    )
    cleaned = [char if char.isalnum() else "_" for char in ascii_text]
    return "_".join(part for part in "".join(cleaned).split("_") if part)


def parse_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value)
    if not text or text == "-":
        return None
    lowered = text.lower()
    if lowered in {"n/c", "nc", "sigiloso", "nao informado", "não informado"}:
        return None
    text = (
        text.replace("R$", "")
        .replace(" ", "")
        .replace(".", "")
        .replace(",", ".")
    )
    text = "".join(char for char in text if char.isdigit() or char in ".-")
    if not text or text == "-":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_bool(value: Any) -> bool | None:
    text = normalize_key_part(value)
    if not text:
        return None
    if text in {"sim", "s", "yes", "true", "1"}:
        return True
    if text in {"nao", "n", "false", "0"}:
        return False
    return None


def parse_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = normalize_text(value)
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def parse_time_value(value: Any) -> time | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    text = normalize_text(value).lower()
    if not text or text in {"n/c", "nc"}:
        return None
    text = text.replace("h", ":")
    if ":" not in text and text.isdigit():
        text = f"{text}:00"
    parts = [part for part in text.split(":") if part != ""]
    try:
        hour = int(parts[0])
        minute = int(parts[1]) if len(parts) > 1 else 0
        return time(hour=hour, minute=minute)
    except (ValueError, IndexError):
        return None


def parse_datetime_pair(date_value: Any, time_value: Any = None) -> datetime | None:
    parsed_date = parse_datetime(date_value)
    parsed_time = parse_time_value(time_value)
    if parsed_date and parsed_time:
        return datetime.combine(parsed_date.date(), parsed_time)
    return parsed_date


def parse_item_number(value: Any, fallback_index: int) -> str:
    if isinstance(value, (datetime, date)):
        return str(value.day)
    text = normalize_text(value)
    if not text or text == "-":
        return str(fallback_index)
    if text.endswith(".0"):
        try:
            return str(int(float(text)))
        except ValueError:
            return text
    return text


def build_tor_id(prefix: str, import_key: str) -> str:
    digest = hashlib.sha1(import_key.encode("utf-8")).hexdigest()[:8].upper()
    return f"TOR-{prefix}-{digest}"


def get_value(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        canonical = canonical_header(key)
        if canonical in row:
            return row.get(canonical)
    return None


def has_meaningful_value(value: Any) -> bool:
    text = normalize_text(value)
    return bool(text and text.lower() not in {"n/c", "nc", "-", "sigiloso"})


def should_import_analyzed_row(row: dict[str, Any]) -> bool:
    selected = normalize_key_part(get_value(row, "selecionado"))
    return selected not in {"nao", "não", "no", "false", "0"}


def ensure_context(db) -> ImportContext:
    from app.auth.models import Tenant, User
    from app.auth.security import hash_password

    primary_user = db.query(User).filter(User.email == PRIMARY_EMAIL).first()
    if primary_user:
        return ImportContext(tenant=primary_user.tenant, user=primary_user)

    fallback_user = db.query(User).filter(User.email == FALLBACK_EMAIL).first()
    if fallback_user:
        tenant = fallback_user.tenant
        primary_user = User(
            email=PRIMARY_EMAIL,
            full_name=DEFAULT_USER_NAME,
            hashed_password=hash_password(DEFAULT_PASSWORD),
            role="admin",
            tenant_id=tenant.id,
        )
        db.add(primary_user)
        db.flush()
        return ImportContext(tenant=tenant, user=primary_user)

    tenant = db.query(Tenant).filter(Tenant.slug == DEFAULT_TENANT_SLUG).first()
    if tenant is None:
        tenant = Tenant(slug=DEFAULT_TENANT_SLUG, name=DEFAULT_TENANT_NAME)
        db.add(tenant)
        db.flush()

    primary_user = User(
        email=PRIMARY_EMAIL,
        full_name=DEFAULT_USER_NAME,
        hashed_password=hash_password(DEFAULT_PASSWORD),
        role="admin",
        tenant_id=tenant.id,
    )
    db.add(primary_user)
    db.flush()
    return ImportContext(tenant=tenant, user=primary_user)


def ensure_portal(db, tenant_id: int, name: str | None, url: str | None = None, created_by: int | None = None) -> CrmPortal | None:
    from app.crm.models import CrmPortal

    portal_name = optional_meaningful_text(name)
    if not portal_name:
        return None
    portal = (
        db.query(CrmPortal)
        .filter(CrmPortal.tenant_id == tenant_id, CrmPortal.name == portal_name)
        .first()
    )
    if portal is None:
        portal = CrmPortal(
            tenant_id=tenant_id,
            name=portal_name,
            url=optional_meaningful_text(url),
            created_by=created_by,
        )
        db.add(portal)
        db.flush()
    elif not portal.url and url:
        portal.url = optional_meaningful_text(url)
    return portal


def ensure_organ(
    db,
    tenant_id: int,
    name: str | None,
    city: str | None = None,
    state: str | None = None,
    created_by: int | None = None,
) -> CrmOrgan | None:
    from app.crm.models import CrmOrgan

    organ_name = optional_text(name)
    if not organ_name:
        return None
    organ = (
        db.query(CrmOrgan)
        .filter(CrmOrgan.tenant_id == tenant_id, CrmOrgan.name == organ_name)
        .first()
    )
    if organ is None:
        organ = CrmOrgan(
            tenant_id=tenant_id,
            name=organ_name,
            city=optional_text(city),
            state=optional_text(state),
            created_by=created_by,
        )
        db.add(organ)
        db.flush()
    else:
        if not organ.city and city:
            organ.city = optional_text(city)
        if not organ.state and state:
            organ.state = optional_text(state)
    return organ


def read_sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [canonical_header(value) for value in rows[0]]
    parsed_rows: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        if not any(value not in (None, "") for value in raw_row):
            continue
        row = {headers[index]: raw_row[index] for index in range(len(headers)) if headers[index]}
        parsed_rows.append(row)
    return parsed_rows


def group_sheet_rows(sheet_name: str, rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if sheet_name == "Processos ganhos":
            key_parts = [
                sheet_name,
                normalize_key_part(row.get("data_da_sessao")),
                normalize_key_part(row.get("orgao")),
                normalize_key_part(row.get("link_pasta_processo")),
                normalize_key_part(row.get("proposta")),
            ]
        else:
            key_parts = [
                sheet_name,
                normalize_key_part(row.get("data_da_sessao")),
                normalize_key_part(row.get("portal")),
                normalize_key_part(row.get("orgao")),
            ]
        groups["|".join(key_parts)].append(row)
    return groups


def upsert_notice_from_group(db, context: ImportContext, import_key: str, rows: list[dict[str, Any]], sheet_name: str) -> CrmNotice:
    from app.crm.models import CrmNotice
    from app.services.crm_notice_sync import sync_notice_relationships

    first = rows[0]
    prefix = "GAN" if sheet_name == "Processos ganhos" else "MON"
    tor_id = build_tor_id(prefix, import_key)
    auction_date = parse_datetime(first.get("data_da_sessao"))

    if sheet_name == "Processos ganhos":
        municipality = optional_text(first.get("endereco"))
        state = optional_text(first.get("uf"))
        organ = ensure_organ(
            db,
            context.tenant.id,
            first.get("orgao"),
            city=municipality,
            state=state,
            created_by=context.user.id,
        )
        notice_fields = {
            "number": tor_id,
            "tor_id": tor_id,
            "municipality_name": municipality,
            "title": None,
            "organ_id": organ.id if organ else None,
            "portal_id": None,
            "auction_date": auction_date,
            "drive_link": optional_text(first.get("link_pasta_processo")),
            "proposal_link": optional_text(first.get("proposta")),
            "supplier_proposal_link": optional_text(first.get("proposta_fornecedor")),
            "address": municipality,
            "state": state,
            "sales_status": optional_text(first.get("situacao")) or "ganho",
            "owner_id": context.user.id,
            "created_by": context.user.id,
            "import_key": import_key,
            "stage": "result",
            "outcome": "won",
            "company_position": "ganho",
        }
    else:
        organ = ensure_organ(
            db,
            context.tenant.id,
            first.get("orgao"),
            created_by=context.user.id,
        )
        portal = ensure_portal(
            db,
            context.tenant.id,
            first.get("portal"),
            created_by=context.user.id,
        )
        notice_fields = {
            "number": tor_id,
            "tor_id": tor_id,
            "municipality_name": None,
            "title": None,
            "organ_id": organ.id if organ else None,
            "portal_id": portal.id if portal else None,
            "auction_date": auction_date,
            "drive_link": None,
            "proposal_link": None,
            "supplier_proposal_link": None,
            "address": None,
            "state": None,
            "sales_status": optional_text(first.get("status")) or "monitoramento",
            "owner_id": context.user.id,
            "created_by": context.user.id,
            "import_key": import_key,
            "stage": "auction",
            "outcome": "pending",
            "company_position": optional_text(first.get("posicao_tor")),
        }

    notice = (
        db.query(CrmNotice)
        .filter(CrmNotice.tenant_id == context.tenant.id, CrmNotice.import_key == import_key)
        .first()
    )
    if notice is None:
        notice = CrmNotice(tenant_id=context.tenant.id, **notice_fields)
        db.add(notice)
        db.flush()
    else:
        for field_name, value in notice_fields.items():
            if value is not None or getattr(notice, field_name) is None:
                setattr(notice, field_name, value)

    sync_notice_relationships(db, notice, created_by=context.user.id)
    return notice


def upsert_group_products(db, context: ImportContext, notice: CrmNotice, rows: list[dict[str, Any]], sheet_name: str) -> int:
    from app.crm.models import CrmNoticeProduct
    from app.services.crm_notice_sync import sync_notice_from_product

    imported = 0
    for index, row in enumerate(rows, start=1):
        if sheet_name == "Processos ganhos":
            item_number = parse_item_number(row.get("item"), index)
            product_code = optional_text(row.get("codigo_do_produto"))
            description = optional_text(row.get("descricao")) or product_code or f"Produto do item {item_number}"
            lot = optional_text(row.get("lote"))
            quantity = parse_float(row.get("quantidade")) or 1.0
            cost = parse_float(row.get("custo_unitario_c_ipi")) or parse_float(row.get("custo_unitario_s_ipi"))
            reference_price = parse_float(row.get("preco_minimo_unitario"))
            reference_total = parse_float(row.get("preco_minimo_total"))
            unit_price = None
            notes = optional_text(row.get("observacao"))
            exclusive_epp = parse_bool(row.get("exclusivo_epp"))
        else:
            item_number = str(index)
            product_code = optional_text(row.get("part_number_do_produto"))
            description = optional_text(row.get("descricao_do_produto")) or product_code or f"Produto monitorado {index}"
            lot = None
            quantity = parse_float(row.get("quantidade")) or 1.0
            cost = parse_float(row.get("custo"))
            reference_price = parse_float(row.get("preco"))
            reference_total = parse_float(row.get("preco_minimo_total"))
            unit_price = parse_float(row.get("preco"))
            notes = optional_text(row.get("status"))
            exclusive_epp = None

        product = (
            db.query(CrmNoticeProduct)
            .filter(CrmNoticeProduct.notice_id == notice.id, CrmNoticeProduct.item_number == item_number)
            .first()
        )
        if product is None:
            product = CrmNoticeProduct(
                tenant_id=context.tenant.id,
                notice_id=notice.id,
                item_number=item_number,
                sort_order=index - 1,
                description=description,
                quantity=quantity,
            )
            db.add(product)
            db.flush()

        product.description = description
        product.lot = lot
        product.product_code = product_code
        product.is_exclusive_epp = exclusive_epp
        product.quantity = quantity
        product.cost = cost
        product.unit_price = unit_price
        product.reference_price = reference_price
        product.reference_total_price = reference_total
        product.notes = notes
        product.sort_order = index - 1

        sync_notice_from_product(db, product, created_by=context.user.id)
        imported += 1

    return imported


def detect_analyzed_sheet_rows(path: Path) -> tuple[str, list[dict[str, Any]]] | None:
    workbook = load_workbook(path, data_only=True, read_only=True)
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = {canonical_header(value) for value in rows[0] if value}
        if {"n_interno", "orgao", "item_1"}.issubset(headers):
            return sheet.title, read_sheet_rows(path, sheet.title)
    return None


def build_analyzed_import_key(row: dict[str, Any], row_index: int) -> str:
    n_interno = optional_meaningful_text(get_value(row, "n interno"))
    if n_interno:
        return f"analisados|{n_interno}"
    key_parts = [
        "analisados",
        normalize_key_part(get_value(row, "orgao")),
        normalize_key_part(get_value(row, "nº pregão")),
        normalize_key_part(get_value(row, "data disputa")),
        str(row_index),
    ]
    return "|".join(key_parts)


def normalize_status_label(value: Any) -> str:
    text = normalize_text(value)
    if not text:
        return "analisado"
    if "🟢" in text:
        return "apto"
    if "🟡" in text:
        return "atencao"
    if "🔴" in text:
        return "risco"
    return text


def get_analyzed_portal_value(row: dict[str, Any]) -> Any:
    portal = get_value(row, "portal")
    if has_meaningful_value(portal):
        return portal
    return optional_meaningful_text(get_value(row, "local"))


def build_notice_notes(row: dict[str, Any]) -> str | None:
    fields = [
        ("Resumo switches", get_value(row, "resumo switches")),
        ("Criterio", get_value(row, "criterio")),
        ("Risco identificado", get_value(row, "risco identificado")),
        ("Validade da proposta", get_value(row, "validade da proposta")),
        (
            "Entrega documentacao",
            get_value(row, "momento entrega documentacao habilitacao"),
        ),
        ("Intervalo", get_value(row, "intervalo")),
        ("Exclusividade", get_value(row, "exclusividade me/epp")),
        ("Valor total edital", get_value(row, "vlr total edital")),
    ]
    parts = [
        f"{label}: {normalize_text(value)}"
        for label, value in fields
        if has_meaningful_value(value)
    ]
    return "\n".join(parts) or None


def upsert_notice_document(
    db,
    context: ImportContext,
    notice: CrmNotice,
    name: str,
    category: str,
    notes: Any,
    sort_order: int,
) -> bool:
    from app.crm.models import CrmChecklistStatus, CrmNoticeDocument

    if not has_meaningful_value(notes):
        return False

    existing = (
        db.query(CrmNoticeDocument)
        .filter(
            CrmNoticeDocument.notice_id == notice.id,
            CrmNoticeDocument.name == name,
        )
        .first()
    )
    if existing is None:
        existing = CrmNoticeDocument(
            tenant_id=context.tenant.id,
            notice_id=notice.id,
            name=name,
            category=category,
            status=CrmChecklistStatus.PENDING,
            is_required=True,
            is_specific=True,
            sort_order=sort_order,
        )
        db.add(existing)
        db.flush()

    existing.category = category
    existing.notes = optional_text(notes)
    existing.sort_order = sort_order
    return True


def upsert_analyzed_documents(db, context: ImportContext, notice: CrmNotice, row: dict[str, Any]) -> int:
    document_fields = [
        ("Habilitacao juridica", "habilitacao", "habilitacao juridica"),
        (
            "Habilitacao fiscal, social e trabalhista",
            "habilitacao",
            "habilitacao fiscal social e trabalhista",
        ),
        (
            "Qualificacao economico-financeira",
            "habilitacao",
            "qualificacao economico financeira",
        ),
        ("Qualificacao tecnica", "habilitacao", "qualificacao tecnica"),
    ]
    total = 0
    for index, (name, category, key) in enumerate(document_fields, start=1):
        if upsert_notice_document(
            db,
            context,
            notice,
            name,
            category,
            get_value(row, key),
            index,
        ):
            total += 1
    return total


def upsert_notice_from_analyzed_row(
    db,
    context: ImportContext,
    row: dict[str, Any],
    row_index: int,
) -> CrmNotice:
    from app.crm.models import CrmNotice
    from app.services.crm_notice_sync import sync_notice_relationships

    import_key = build_analyzed_import_key(row, row_index)
    n_interno = optional_text(get_value(row, "n interno"))
    tor_id = n_interno or build_tor_id("ANA", import_key)
    city = optional_meaningful_text(get_value(row, "cidade"))
    state = optional_meaningful_text(get_value(row, "uf"))
    organ = ensure_organ(
        db,
        context.tenant.id,
        get_value(row, "orgao"),
        city=city,
        state=state,
        created_by=context.user.id,
    )
    portal = ensure_portal(
        db,
        context.tenant.id,
        get_analyzed_portal_value(row),
        created_by=context.user.id,
    )
    notice_number = tor_id
    preg_number = optional_meaningful_text(get_value(row, "nº pregão"))
    title_parts = [
        part
        for part in [
            preg_number,
            optional_meaningful_text(get_value(row, "resumo switches")),
        ]
        if part
    ]
    auction_date = parse_datetime_pair(
        get_value(row, "data disputa"),
        get_value(row, "hora disputa"),
    )

    notice_fields = {
        "number": notice_number,
        "tor_id": tor_id,
        "municipality_name": city,
        "title": " - ".join(title_parts) or None,
        "organ_id": organ.id if organ else None,
        "portal_id": portal.id if portal else None,
        "modality": optional_meaningful_text(get_value(row, "tipo licitação")),
        "auction_date": auction_date,
        "estimated_value": parse_float(get_value(row, "valor total switches")),
        "drive_link": optional_meaningful_text(get_value(row, "url drive")),
        "address": optional_meaningful_text(get_value(row, "endereço")),
        "state": state,
        "particularities": build_notice_notes(row),
        "sales_status": normalize_status_label(get_value(row, "status")),
        "owner_id": context.user.id,
        "created_by": context.user.id,
        "import_key": import_key,
        "stage": "triage",
        "outcome": "pending",
        "company_position": optional_meaningful_text(get_value(row, "criterio")),
    }

    notice = (
        db.query(CrmNotice)
        .filter(CrmNotice.tenant_id == context.tenant.id, CrmNotice.import_key == import_key)
        .first()
    )
    if notice is None:
        notice = CrmNotice(tenant_id=context.tenant.id, **notice_fields)
        db.add(notice)
        db.flush()
    else:
        for field_name, value in notice_fields.items():
            if value is not None or getattr(notice, field_name) is None:
                setattr(notice, field_name, value)

    sync_notice_relationships(db, notice, created_by=context.user.id)
    return notice


def iter_analyzed_items(row: dict[str, Any]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for index in range(1, 11):
        raw_description = get_value(row, f"item {index}")
        if not has_meaningful_value(raw_description):
            continue
        description = normalize_text(raw_description)
        item_number = parse_item_number(
            get_value(row, f"número item edital {index}"),
            index,
        )
        quantity = parse_float(get_value(row, f"quantidade item {index}")) or 1.0
        reference_price = parse_float(get_value(row, f"preço item {index}"))
        reference_total = (
            round(reference_price * quantity, 4)
            if reference_price is not None and quantity is not None
            else None
        )
        items.append(
            {
                "item_number": item_number,
                "lot": optional_meaningful_text(get_value(row, f"lote item {index}")),
                "description": description,
                "quantity": quantity,
                "reference_price": reference_price,
                "reference_total": reference_total,
                "exclusive_epp": parse_bool(
                    get_value(row, f"exclusividade me/epp item {index}")
                ),
                "sort_order": index - 1,
            }
        )
    return items


def upsert_analyzed_products(
    db,
    context: ImportContext,
    notice: CrmNotice,
    row: dict[str, Any],
) -> int:
    from app.crm.models import CrmNoticeProduct
    from app.services.crm_notice_sync import sync_notice_from_product

    imported = 0
    for item in iter_analyzed_items(row):
        product = (
            db.query(CrmNoticeProduct)
            .filter(
                CrmNoticeProduct.notice_id == notice.id,
                CrmNoticeProduct.item_number == item["item_number"],
            )
            .first()
        )
        if product is None:
            product = CrmNoticeProduct(
                tenant_id=context.tenant.id,
                notice_id=notice.id,
                item_number=item["item_number"],
                sort_order=item["sort_order"],
                description=item["description"],
                quantity=item["quantity"],
            )
            db.add(product)
            db.flush()

        product.description = item["description"]
        product.lot = item["lot"]
        product.is_exclusive_epp = item["exclusive_epp"]
        product.quantity = item["quantity"]
        product.reference_price = item["reference_price"]
        product.reference_total_price = item["reference_total"]
        product.unit_price = item["reference_price"]
        product.notes = optional_meaningful_text(get_value(row, "risco identificado"))
        product.sort_order = item["sort_order"]

        sync_notice_from_product(db, product, created_by=context.user.id)
        imported += 1
    return imported


def run_analyzed_import(
    db,
    context: ImportContext,
    rows: list[dict[str, Any]],
) -> dict[str, Any]:
    total_notices = 0
    total_products = 0
    total_documents = 0
    skipped_rows = 0

    for row_index, row in enumerate(rows, start=2):
        if not should_import_analyzed_row(row):
            skipped_rows += 1
            continue
        if not has_meaningful_value(get_value(row, "n interno")) and not has_meaningful_value(get_value(row, "orgao")):
            skipped_rows += 1
            continue

        notice = upsert_notice_from_analyzed_row(db, context, row, row_index)
        total_notices += 1
        total_products += upsert_analyzed_products(db, context, notice, row)
        total_documents += upsert_analyzed_documents(db, context, notice, row)

    return {
        "grupos_processados": total_notices,
        "itens_processados": total_products,
        "documentos_processados": total_documents,
        "linhas_ignoradas": skipped_rows,
        "formato": "planilha_analisada",
    }


def build_import_context_for_user(user: User) -> ImportContext:
    tenant = user.tenant
    if tenant is None:
        raise ValueError("Usuario importador sem tenant associado.")
    return ImportContext(tenant=tenant, user=user)


def run_import(path: Path, *, context_override: ImportContext | None = None) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada em {path}")

    from app.db.init_db import init_db
    from app.db.session import SessionLocal

    db = SessionLocal()
    try:
        init_db(db)
        context = context_override or ensure_context(db)
        analyzed_sheet = detect_analyzed_sheet_rows(path)
        if analyzed_sheet is not None:
            sheet_name, rows = analyzed_sheet
            summary = run_analyzed_import(db, context, rows)
            summary.update(
                {
                    "tenant": context.tenant.slug,
                    "user": context.user.email,
                    "sheet": sheet_name,
                }
            )
            db.commit()
            print("Importacao concluida:", summary)
            return summary

        sheet_names = ["Processos ganhos", "Processos - Monitoramento"]
        total_notices = 0
        total_products = 0

        for sheet_name in sheet_names:
            rows = read_sheet_rows(path, sheet_name)
            for import_key, group_rows in group_sheet_rows(sheet_name, rows).items():
                notice = upsert_notice_from_group(db, context, import_key, group_rows, sheet_name)
                total_notices += 1
                total_products += upsert_group_products(db, context, notice, group_rows, sheet_name)

        db.commit()
        summary = {
            "tenant": context.tenant.slug,
            "user": context.user.email,
            "grupos_processados": total_notices,
            "itens_processados": total_products,
        }
        print("Importacao concluida:", summary)
        return summary
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
