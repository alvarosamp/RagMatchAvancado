from __future__ import annotations

import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.crm.models import CrmCatalogProduct


BLOCKED_AVAILABILITY_VALUES = {
    "",
    "-",
    "nao vender",
    "xwdm nao vender",
}


@dataclass
class LpuImportSummary:
    sheets: int = 0
    processed: int = 0
    created: int = 0
    updated: int = 0
    skipped: int = 0
    skipped_duplicates: int = 0
    duplicate_updates: int = 0
    removed_stale: int = 0
    lpu_version: str = ""
    lpu_drive_url: str = ""

    def as_dict(self) -> dict[str, Any]:
        return {
            "sheets": self.sheets,
            "items": self.processed,
            "total_items": self.processed,
            "processed": self.processed,
            "created": self.created,
            "updated": self.updated,
            "skipped": self.skipped,
            "skipped_duplicates": self.skipped_duplicates,
            "duplicate_updates": self.duplicate_updates,
            "removed_stale": self.removed_stale,
            "lpu_version": self.lpu_version,
            "lpu_drive_url": self.lpu_drive_url,
        }


def import_lpu_catalog(
    path: Path,
    *,
    db: Session,
    tenant_id: int,
    user_id: int | None,
    lpu_drive_url: str,
) -> dict[str, Any]:
    lpu_drive_url = _normalize_lpu_drive_url(lpu_drive_url)
    workbook = load_workbook(path, data_only=True)
    lpu_version = _build_lpu_version(path)
    summary = LpuImportSummary(lpu_version=lpu_version, lpu_drive_url=lpu_drive_url)
    seen_skus: set[str] = set()
    imported_categories: set[str] = set()
    imported_skus: set[str] = set()
    has_proposal_sheet = any(_norm(sheet.title) == "proposta" for sheet in workbook.worksheets)

    for sheet in workbook.worksheets:
        if has_proposal_sheet and _norm(sheet.title) == "fob":
            continue
        header_row, header = _find_header(sheet)
        if not any(header):
            continue
        summary.sheets += 1
        normalized_header = [_norm(value) for value in header]

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            if not _row_has_values(row):
                continue
            record = _row_to_record(header, normalized_header, row, sheet.title)
            if not record:
                summary.skipped += 1
                continue
            if record["sku"] in seen_skus:
                summary.duplicate_updates += 1
                summary.skipped_duplicates += 1
            else:
                seen_skus.add(record["sku"])

            summary.processed += 1
            record["lpu_version"] = lpu_version
            record["lpu_drive_url"] = lpu_drive_url
            imported_categories.add(record["category"])
            imported_skus.add(record["sku"])
            existing = (
                db.query(CrmCatalogProduct)
                .filter(
                    CrmCatalogProduct.tenant_id == tenant_id,
                    CrmCatalogProduct.sku == record["sku"],
                )
                .first()
            )
            if existing:
                for key, value in record.items():
                    setattr(existing, key, value)
                db.add(existing)
                summary.updated += 1
            else:
                db.add(
                    CrmCatalogProduct(
                        tenant_id=tenant_id,
                        created_by=user_id,
                        **record,
                    )
                )
                summary.created += 1
            db.flush()

    if imported_skus:
        stale_rows = (
            db.query(CrmCatalogProduct)
            .filter(
                CrmCatalogProduct.tenant_id == tenant_id,
                CrmCatalogProduct.category.in_(imported_categories),
                CrmCatalogProduct.sku.notin_(imported_skus),
            )
            .all()
        )
        summary.removed_stale = len(stale_rows)
        for row in stale_rows:
            db.delete(row)

    db.commit()
    return summary.as_dict()


def _build_lpu_version(path: Path) -> str:
    stem = Path(path).stem if path else "lpu"
    safe_stem = _safe_sku(stem.lower()) or "lpu"
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    return f"{safe_stem}-{stamp}"


def _normalize_lpu_drive_url(value: str | None) -> str:
    url = str(value or "").strip()
    if not url:
        raise ValueError("Informe o link do Drive da LPU antes de importar.")
    lower_url = url.lower()
    if not lower_url.startswith(("https://", "http://")):
        raise ValueError("O link do Drive da LPU precisa comecar com http:// ou https://.")
    if "drive.google.com" not in lower_url and "docs.google.com" not in lower_url:
        raise ValueError("Informe um link valido do Google Drive ou Google Docs para a LPU.")
    return url


def _row_to_record(
    header: list[str],
    normalized_header: list[str],
    row: tuple[Any, ...],
    sheet_name: str,
) -> dict[str, Any] | None:
    values = {normalized_header[index]: row[index] for index in range(min(len(row), len(header)))}
    proposal_record = _proposal_row_to_record(values, sheet_name)
    if proposal_record:
        return proposal_record

    brand = _clean_text(_get(values, "marca")) or "TOR"
    pn_tor = _clean_text(_get(values, "pn tor", "part number tor", "partnumber tor"))
    if _is_blocked_pn(pn_tor):
        return None
    description_value = _clean_text(_get(values, "description", "descricao"))
    manufacturer_part_number = _clean_text(_get(values, "part no", "part number", "partno", "part no mfr"))
    availability = _clean_text(_get(values, "disponibilidade"))
    if _is_unavailable(availability):
        return None
    price = _to_float(_get(values, "preco", "preço", "price"))
    cost = _to_float(_get(values, "custo final", "custo", "preco", "preco final", "preço", "preço final", "price"))
    minimum_price = _to_float_or_none(_get(values, "preco minimo", "preco mínimo", "preço minimo", "preço mínimo"))
    if minimum_price is None:
        minimum_price = price
    category = _category_from_sheet(sheet_name)

    if category == "switch":
        model = description_value or pn_tor
        description = _clean_text(_get(values, "vig")) or model
    else:
        description = description_value
        model = pn_tor or description

    sku = _build_sku(
        pn_tor=pn_tor,
        part_no=manufacturer_part_number,
        model=model,
        category=category,
    )
    if not sku or not description:
        return None

    name = " ".join(part for part in [brand, model] if part) or description
    keywords = " ".join(
        part
        for part in [
            brand,
            pn_tor,
            manufacturer_part_number,
            model,
            description,
            _clean_text(_get(values, "rate")),
            _clean_text(_get(values, "distance")),
        ]
        if part
    )
    notes = f"Disponibilidade: {availability}" if availability else ""

    return {
        "name": name[:255],
        "description": description,
        "category": category,
        "brand": brand,
        "model": (model or sku)[:255],
        "manufacturer_part_number": manufacturer_part_number[:255] or None,
        "specification": description,
        "sku": sku[:255],
        "keywords": keywords,
        "unit": "UN",
        "cost": cost,
        "min_price": minimum_price if minimum_price is not None else cost,
        "tax_percent": 0.0,
        "margin_percent": 0.0,
        "notes": notes,
        "is_active": True,
    }


def _proposal_row_to_record(values: dict[str, Any], sheet_name: str) -> dict[str, Any] | None:
    item_number = _clean_text(_get(values, "item"))
    model = _clean_text(_get(values, "modelo"))
    minimum_price = _to_float_or_none(_get(values, "preco minimo", "preco mínimo", "preço minimo", "preço mínimo"))
    if not item_number or not (model or minimum_price is not None):
        return None

    brand = _clean_text(_get(values, "marca")) or "TOR"
    description = _clean_text(
        _get(
            values,
            "description",
            "descricao",
            "descrição",
            "descricao conforme descricao complementar no termo de referencia deste processo",
        )
    )
    model = model or description or item_number
    description = description or model
    proposal_price = _to_float_or_none(_get(values, "preco proposta", "preço proposta", "preco", "preço", "price"))
    price = minimum_price if minimum_price is not None else proposal_price or 0.0
    sku = _safe_sku(f"{item_number}-{model}") or _safe_sku(model)
    if not sku or not description:
        return None

    notes = _clean_text(_get(values, "disponibilidade"))
    if _is_unavailable(notes):
        return None
    if proposal_price is not None:
        notes = "\n".join(part for part in [notes, f"Preco proposta: {proposal_price:.4f}"] if part)
    if notes and "Preco proposta:" not in notes:
        notes = f"Disponibilidade: {notes}"

    category = _category_from_sheet(sheet_name)
    unit = _clean_text(_get(values, "unidade", "unidade ")) or "UN"
    name = " ".join(part for part in [brand, model] if part) or description
    keywords = " ".join(part for part in [item_number, brand, model, description] if part)

    return {
        "name": name[:255],
        "description": description,
        "category": category,
        "brand": brand,
        "model": model[:255],
        "specification": description,
        "sku": sku[:255],
        "keywords": keywords,
        "unit": unit[:40],
        "cost": price,
        "min_price": minimum_price if minimum_price is not None else price,
        "tax_percent": 0.0,
        "margin_percent": 0.0,
        "notes": notes,
        "is_active": True,
    }


def _find_header(sheet) -> tuple[int, list[str]]:
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), values_only=True), start=1):
        header = [str(value or "").strip() for value in row]
        normalized = {_norm(value) for value in header if value}
        if {"marca", "modelo"}.issubset(normalized) or {"item", "marca"}.issubset(normalized):
            return row_index, header
    first = next(sheet.iter_rows(max_row=1, values_only=True))
    return 1, [str(value or "").strip() for value in first]


def _get(values: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        normalized = _norm(key)
        if normalized in values:
            return values[normalized]
    return None


def _norm(value: Any) -> str:
    text = str(value or "").strip().lower()
    ascii_text = (
        unicodedata.normalize("NFKD", text)
        .encode("ascii", "ignore")
        .decode("ascii")
    )
    cleaned = "".join(char if char.isalnum() else " " for char in ascii_text)
    return " ".join(cleaned.split())


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _row_has_values(row: tuple[Any, ...]) -> bool:
    return any(_clean_text(value) for value in row)


def _valid_sku(value: str) -> str:
    if _is_blocked_pn(value) or value in {"0", "0.0"}:
        return ""
    return value


def _build_sku(*, pn_tor: str, part_no: str, model: str, category: str) -> str:
    pn = _valid_sku(pn_tor)
    part = _valid_sku(part_no)
    model_value = _valid_sku(model)
    if category == "switch" and pn and model_value and pn != model_value:
        return _safe_sku(f"{pn}-{model_value}")
    return _safe_sku(pn or part or model_value)


def _safe_sku(value: str) -> str:
    safe = "".join(ch if ch.isalnum() else "-" for ch in value.strip())
    return "-".join(part for part in safe.split("-") if part)


def _dedupe_sku(sku: str, model: Any, seen_skus: set[str]) -> str:
    if sku not in seen_skus:
        return sku
    model_sku = _safe_sku(str(model or ""))
    if model_sku and model_sku not in seen_skus:
        return model_sku
    base = sku
    counter = 2
    while f"{base}-{counter}" in seen_skus:
        counter += 1
    return f"{base}-{counter}"


def _to_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _to_float_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("R$", "").replace(" ", "")
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _category_from_sheet(sheet_name: str) -> str:
    normalized = _norm(sheet_name)
    if "switch" in normalized:
        return "switch"
    if "cabo" in normalized:
        return "cabo"
    if "transceiver" in normalized:
        return "transceiver"
    return normalized or "produto"


def _is_unavailable(value: str) -> bool:
    normalized = _norm(value)
    if normalized in BLOCKED_AVAILABILITY_VALUES:
        return True
    return "nao vender" in normalized


def _is_blocked_pn(value: Any) -> bool:
    normalized = _norm(value)
    if not normalized:
        return True
    return normalized in {
        "-",
        "xwdm nao vender",
    }
