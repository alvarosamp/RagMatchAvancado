"""
services/export_service.py
───────────────────────────
Gera relatórios de matching em três formatos:
  - XLSX  →  planilha com aba resumo + aba detalhada
  - PDF   →  relatório formatado pronto para licitação
  - CSV   →  dados brutos para integração externa

Uso:
    from app.services.export_service import export_xlsx, export_pdf, export_csv
"""

from __future__ import annotations

import csv
import io
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)

from app.logs.config import logger

# ── Cores padrão ─────────────────────────────────────────────────────────────
COR_VERDE      = "00A86B"
COR_VERMELHO   = "D32F2F"
COR_AMARELO    = "F9A825"
COR_AZUL_ESCURO = "1565C0"
COR_CINZA      = "F5F5F5"
COR_CABECALHO  = "1A237E"


# ─────────────────────────────────────────────────────────────────────────────
# XLSX
# ─────────────────────────────────────────────────────────────────────────────

def export_xlsx(data: dict) -> bytes:
    """
    Gera planilha Excel com:
    - Aba 'Resumo'   → ranking de produtos por score
    - Aba 'Detalhes' → resultado por produto × requisito

    Args:
        data: dict retornado por POST /editais/{id}/match

    Returns:
        bytes do arquivo .xlsx
    """
    wb = Workbook()
    _build_resumo_sheet(wb, data)
    _build_detalhes_sheet(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info(f"[Export] XLSX gerado — {len(data.get('results', []))} produtos")
    return buf.getvalue()


def _build_resumo_sheet(wb: Workbook, data: dict):
    ws = wb.active
    ws.title = "Resumo"

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = "RELATÓRIO DE MATCHING — EDITAL MATCHER"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", start_color=COR_CABECALHO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws["D2"] = f"Edital ID: {data.get('edital_id', '-')}"
    ws["D2"].font = Font(italic=True, size=9, color="666666")

    # ── Cabeçalho da tabela ───────────────────────────────────────────────────
    headers = ["#", "Modelo", "Score Geral", "Status", "Atende", "Não Atende"]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", start_color=COR_AZUL_ESCURO)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = _thin_border()

    # ── Dados ─────────────────────────────────────────────────────────────────
    results = data.get("results", [])
    for i, r in enumerate(results, 1):
        row = header_row + i
        score   = r.get("overall_score", 0)
        status  = r.get("status", "")
        details = r.get("details", [])

        atende     = sum(1 for d in details if d.get("status") == "atende")
        nao_atende = sum(1 for d in details if d.get("status") == "nao_atende")

        values = [i, r.get("model", ""), f"{score:.0%}", _status_label(status), atende, nao_atende]
        fill_color = _row_color(score)

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill      = PatternFill("solid", start_color=fill_color)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = _thin_border()
            if col == 3:
                cell.font = Font(bold=True)

    # ── Largura das colunas ───────────────────────────────────────────────────
    for col, w in zip("ABCDEF", [5, 20, 14, 16, 10, 12]):
        ws.column_dimensions[col].width = w


def _build_detalhes_sheet(wb: Workbook, data: dict):
    ws = wb.create_sheet("Detalhes")

    headers = ["Modelo", "Requisito", "Exigido", "Encontrado", "Score", "Status", "Justificativa LLM"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", start_color=COR_CABECALHO)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = _thin_border()

    row = 2
    for r in data.get("results", []):
        model = r.get("model", "")
        for d in r.get("details", []):
            score  = d.get("final_score", 0)
            values = [
                model,
                d.get("attribute", ""),
                d.get("required", ""),
                d.get("found", ""),
                f"{score:.0%}",
                _status_label(d.get("status", "")),
                d.get("reasoning", ""),
            ]
            fill = _row_color(score)
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill      = PatternFill("solid", start_color=fill)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border    = _thin_border()
            row += 1

    for col, w in zip("ABCDEFG", [18, 22, 22, 22, 8, 14, 45]):
        ws.column_dimensions[col].width = w


# ─────────────────────────────────────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────────────────────────────────────

def export_pdf(data: dict) -> bytes:
    """
    Gera relatório PDF formatado com:
    - Capa / cabeçalho
    - Tabela de ranking
    - Detalhes por produto (top 5)

    Returns:
        bytes do arquivo .pdf
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles  = getSampleStyleSheet()
    story   = []
    results = data.get("results", [])

    # ── Cabeçalho ─────────────────────────────────────────────────────────────
    title_style = ParagraphStyle(
        "titulo", parent=styles["Title"],
        fontSize=18, textColor=colors.HexColor(f"#{COR_CABECALHO}"),
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.gray, spaceAfter=12,
    )

    story.append(Paragraph("Relatório de Matching de Editais", title_style))
    story.append(Paragraph(
        f"Edital ID: {data.get('edital_id', '-')} &nbsp;|&nbsp; "
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} &nbsp;|&nbsp; "
        f"{len(results)} produto(s) avaliado(s)",
        sub_style,
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor(f"#{COR_CABECALHO}")))
    story.append(Spacer(1, 0.4*cm))

    # ── Tabela ranking ─────────────────────────────────────────────────────────
    story.append(Paragraph("Ranking de Produtos", styles["Heading2"]))
    story.append(Spacer(1, 0.2*cm))

    table_data = [["#", "Modelo", "Score", "Status", "Resumo"]]
    for i, r in enumerate(results, 1):
        score = r.get("overall_score", 0)
        table_data.append([
            str(i),
            r.get("model", ""),
            f"{score:.0%}",
            _status_label(r.get("status", "")),
            r.get("summary", "")[:80],
        ])

    t = Table(table_data, colWidths=[1.2*cm, 4*cm, 2*cm, 3.5*cm, 6.5*cm])
    t.setStyle(TableStyle([
        # Cabeçalho
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor(f"#{COR_CABECALHO}")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 9),
        ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
        # Dados
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("ALIGN",       (0, 1), (3,  -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [
            colors.HexColor("#F8F9FA"), colors.white
        ]),
        # Grid
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0,0), (-1, -1), 4),
    ]))

    # Colorir coluna score por valor
    for i, r in enumerate(results, 1):
        score = r.get("overall_score", 0)
        bg = colors.HexColor(f"#{_row_color(score)}")
        t.setStyle(TableStyle([("BACKGROUND", (2, i), (2, i), bg)]))

    story.append(t)
    story.append(Spacer(1, 0.6*cm))

    # ── Detalhes top 5 ────────────────────────────────────────────────────────
    story.append(Paragraph("Detalhes por Produto (Top 5)", styles["Heading2"]))

    for r in results[:5]:
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph(
            f"<b>{r.get('model', '')}</b> — Score: {r.get('overall_score', 0):.0%}  |  {r.get('summary', '')}",
            styles["Normal"],
        ))
        story.append(Spacer(1, 0.15*cm))

        det_data = [["Requisito", "Exigido", "Encontrado", "Score", "Status"]]
        for d in r.get("details", []):
            det_data.append([
                d.get("attribute", ""),
                d.get("required", "")[:30],
                d.get("found", "")[:30],
                f"{d.get('final_score', 0):.0%}",
                _status_label(d.get("status", "")),
            ])

        dt = Table(det_data, colWidths=[4*cm, 4*cm, 3.5*cm, 1.8*cm, 3.4*cm])
        dt.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#455A64")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 7.5),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FAFAFA"), colors.white]),
            ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",  (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",(0,0), (-1, -1), 3),
        ]))
        story.append(dt)

    doc.build(story)
    buf.seek(0)
    logger.info(f"[Export] PDF gerado — {len(results)} produtos")
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# CSV
# ─────────────────────────────────────────────────────────────────────────────

def export_csv(data: dict) -> bytes:
    """
    Gera CSV com todos os resultados linha a linha (produto × requisito).

    Returns:
        bytes UTF-8 com BOM (compatível com Excel BR)
    """
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")

    writer.writerow([
        "edital_id", "modelo", "score_geral", "status_geral",
        "requisito", "exigido", "encontrado",
        "score_item", "status_item", "justificativa_llm",
    ])

    edital_id = data.get("edital_id", "")
    for r in data.get("results", []):
        model        = r.get("model", "")
        score_geral  = f"{r.get('overall_score', 0):.4f}"
        status_geral = _status_label(r.get("status", ""))

        for d in r.get("details", []):
            writer.writerow([
                edital_id,
                model,
                score_geral,
                status_geral,
                d.get("attribute", ""),
                d.get("required", ""),
                d.get("found", ""),
                f"{d.get('final_score', 0):.4f}",
                _status_label(d.get("status", "")),
                d.get("reasoning", "").replace("\n", " "),
            ])

    logger.info(f"[Export] CSV gerado — {len(data.get('results', []))} produtos")
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _status_label(status: str) -> str:
    MAP = {
        "atende":     "✅ Atende",
        "nao_atende": "❌ Não Atende",
        "verificar":  "⚠️ Verificar",
    }
    return MAP.get(status.lower() if status else "", status)


def _row_color(score: float) -> str:
    """Retorna cor de fundo hex baseada no score."""
    if score >= 0.75:
        return "C8E6C9"   # verde claro
    elif score >= 0.45:
        return "FFF9C4"   # amarelo claro
    return "FFCDD2"       # vermelho claro


def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)