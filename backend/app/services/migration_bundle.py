"""Create, validate and restore auditable migration bundles.

The PostgreSQL custom dump is the authoritative backup.  XLSX files are a
human-readable audit layer and must never be used as the sole disaster-recovery
copy of the platform.
"""

from __future__ import annotations

import hashlib
import json
import os
import shutil
import subprocess
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import MetaData, Table, func, inspect, select

from app.db.session import engine
from app.logs.config import logger
from app.services.object_storage import get_bytes, object_storage_enabled, put_bytes


BUNDLE_VERSION = 1
AUDIT_TABLES = (
    "tenants", "users", "products", "editais", "requirements", "matching_results",
    "document_files", "analysis_documents", "analysis_items", "jobs",
)
EXCLUDED_COLUMNS = {"users": {"hashed_password"}}


def create_bundle(output_dir: Path, *, include_files: bool = True) -> dict[str, Any]:
    """Create a portable package without changing database records."""
    output_dir = output_dir.resolve()
    if output_dir.exists() and any(output_dir.iterdir()):
        raise ValueError(f"Diretorio de destino nao esta vazio: {output_dir}")
    output_dir.mkdir(parents=True, exist_ok=True)

    _dump_database(output_dir / "database.dump")
    table_counts = _table_counts()
    _write_audit_workbook(output_dir / "auditoria.xlsx", table_counts)
    copied_files = _copy_file_roots(output_dir / "files") if include_files else []
    object_files = _export_object_storage(output_dir / "objects") if include_files else []

    manifest = {
        "bundle_version": BUNDLE_VERSION,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "database_url_redacted": _redact_database_url(os.environ.get("DATABASE_URL", "")),
        "table_counts": table_counts,
        "files": _file_inventory(output_dir),
        "copied_file_count": len(copied_files),
        "object_file_count": len(object_files),
    }
    (output_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def validate_bundle(bundle_dir: Path, *, compare_database: bool = False) -> dict[str, Any]:
    """Validate integrity before restoration; this function never writes data."""
    bundle_dir = bundle_dir.resolve()
    manifest_path = bundle_dir / "manifest.json"
    if not manifest_path.exists():
        raise ValueError("manifest.json nao encontrado no pacote.")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if manifest.get("bundle_version") != BUNDLE_VERSION:
        raise ValueError(f"Versao de pacote nao suportada: {manifest.get('bundle_version')}")
    if not (bundle_dir / "database.dump").is_file():
        raise ValueError("database.dump nao encontrado no pacote.")

    failures: list[str] = []
    for entry in manifest.get("files", []):
        path = bundle_dir / entry["path"]
        if not path.is_file():
            failures.append(f"Arquivo ausente: {entry['path']}")
        elif _sha256(path) != entry["sha256"]:
            failures.append(f"Hash divergente: {entry['path']}")

    result: dict[str, Any] = {"valid": not failures, "failures": failures, "manifest": manifest}
    if compare_database:
        current = _table_counts()
        expected = manifest.get("table_counts", {})
        result["database_counts"] = {name: {"expected": count, "actual": current.get(name)} for name, count in expected.items()}
        result["database_matches"] = all(current.get(name) == count for name, count in expected.items())
    return result


def restore_bundle(bundle_dir: Path, *, confirm_restore: bool, restore_files: bool = True) -> dict[str, Any]:
    """Restore a verified bundle. Explicit confirmation prevents accidental overwrite."""
    if not confirm_restore:
        raise ValueError("Restauracao bloqueada. Informe --confirm-restore explicitamente.")
    validation = validate_bundle(bundle_dir)
    if not validation["valid"]:
        raise ValueError("Pacote invalido: " + "; ".join(validation["failures"]))

    dump = bundle_dir.resolve() / "database.dump"
    database_url = _pg_cli_url(os.environ.get("DATABASE_URL", ""))
    subprocess.run(
        ["pg_restore", "--clean", "--if-exists", "--no-owner", "--no-privileges", "--dbname", database_url, str(dump)],
        check=True,
    )
    restored_files = _restore_file_roots(bundle_dir / "files") if restore_files else 0
    restored_objects = _restore_object_storage(bundle_dir / "objects") if restore_files else 0
    comparison = validate_bundle(bundle_dir, compare_database=True)
    return {"restored_files": restored_files, "restored_objects": restored_objects, "database_matches": comparison.get("database_matches", False)}


def _dump_database(path: Path) -> None:
    database_url = _pg_cli_url(os.environ.get("DATABASE_URL", ""))
    if not database_url:
        raise ValueError("DATABASE_URL nao configurada para gerar o dump.")
    subprocess.run(
        ["pg_dump", "--format=custom", "--no-owner", "--no-privileges", "--file", str(path), database_url],
        check=True,
    )


def _table_counts() -> dict[str, int]:
    inspector = inspect(engine)
    metadata = MetaData()
    counts: dict[str, int] = {}
    with engine.connect() as connection:
        for name in sorted(inspector.get_table_names()):
            table = Table(name, metadata, autoload_with=engine)
            counts[name] = int(connection.execute(select(func.count()).select_from(table)).scalar_one())
    return counts


def _write_audit_workbook(path: Path, table_counts: dict[str, int]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    summary.append(["Pacote de migração", datetime.now(timezone.utc).replace(tzinfo=None)])
    summary.append(["Tabela", "Registros"])
    for table_name, count in table_counts.items():
        summary.append([table_name, count])

    inspector = inspect(engine)
    available = set(inspector.get_table_names())
    metadata = MetaData()
    for table_name in _audit_table_names(available):
        table = Table(table_name, metadata, autoload_with=engine)
        columns = [column.name for column in table.columns if column.name not in EXCLUDED_COLUMNS.get(table_name, set())]
        sheet = workbook.create_sheet(_sheet_name(table_name))
        sheet.append(columns)
        with engine.connect() as connection:
            for row in connection.execute(select(*(table.c[name] for name in columns))).mappings():
                sheet.append([_xlsx_value(row[name]) for name in columns])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        for column_cells in sheet.columns:
            width = min(50, max(12, max(len(str(cell.value or "")) for cell in column_cells) + 2))
            sheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def _copy_file_roots(destination: Path) -> list[str]:
    roots = [Path(value) for value in os.getenv("MIGRATION_FILE_ROOTS", "/data,/documents").split(",") if value.strip()]
    copied: list[str] = []
    for root in roots:
        if not root.exists() or not root.is_dir():
            continue
        target = destination / root.name
        shutil.copytree(root, target, dirs_exist_ok=True, ignore=shutil.ignore_patterns("tmp_uploads", "backups"))
        copied.extend(str(path.relative_to(destination.parent)) for path in target.rglob("*") if path.is_file())
    return copied


def _restore_file_roots(source: Path) -> int:
    if not source.exists():
        return 0
    restored = 0
    for root in source.iterdir():
        target = Path("/") / root.name
        if not root.is_dir():
            continue
        for file_path in root.rglob("*"):
            if not file_path.is_file():
                continue
            destination = target / file_path.relative_to(root)
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(file_path, destination)
            restored += 1
    return restored


def _export_object_storage(destination: Path) -> list[str]:
    # Object storage is optional; current object keys are already recorded in the DB dump.
    if not object_storage_enabled():
        return []
    from app.services.object_storage import list_objects
    copied: list[str] = []
    for item in list_objects():
        key = item["key"]
        target = destination / key
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(get_bytes(key))
        copied.append(key)
    return copied


def _restore_object_storage(source: Path) -> int:
    if not source.exists() or not object_storage_enabled():
        return 0
    restored = 0
    for path in source.rglob("*"):
        if path.is_file():
            key = path.relative_to(source).as_posix()
            put_bytes(key, path.read_bytes(), content_type="application/octet-stream")
            restored += 1
    return restored


def _file_inventory(root: Path) -> list[dict[str, Any]]:
    return [
        {"path": path.relative_to(root).as_posix(), "size_bytes": path.stat().st_size, "sha256": _sha256(path)}
        for path in sorted(root.rglob("*")) if path.is_file()
    ]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _xlsx_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, (datetime, date)):
        return value.replace(tzinfo=None) if isinstance(value, datetime) and value.tzinfo else value
    return value


def _sheet_name(name: str) -> str:
    return {"editais": "Editais", "requirements": "Requisitos", "matching_results": "Resultados", "crm_notices": "CRM Editais"}.get(name, name[:31])


def _audit_table_names(available: set[str]) -> list[str]:
    names = [name for name in AUDIT_TABLES if name in available]
    names.extend(sorted(name for name in available if name.startswith("crm_") and name not in names))
    return names


def _pg_cli_url(url: str) -> str:
    return url.replace("postgresql+psycopg2://", "postgresql://", 1)


def _redact_database_url(url: str) -> str:
    if "@" not in url:
        return url
    return "postgresql://***:***@" + url.split("@", 1)[1]
