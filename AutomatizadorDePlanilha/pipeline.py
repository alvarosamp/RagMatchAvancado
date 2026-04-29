from __future__ import annotations

import os
import re
import sys
import json
import time
import argparse
import warnings
import unicodedata
import multiprocessing as mp
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook

warnings.filterwarnings(
    "ignore",
    message=r"Field .* has conflict with protected namespace \"model_\".*",
    category=UserWarning,
)


# =========================================================
# IMPORTS LOCAIS
# =========================================================

def _ensure_backend_on_syspath() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    backend_dir = repo_root / "backend"
    backend_str = str(backend_dir)
    if backend_str not in sys.path:
        sys.path.insert(0, backend_str)


_ensure_backend_on_syspath()

from app.pipeline.docling_parser import ParsedDocument, parse_pdf, release_converter  # type: ignore
from extrator_evidencia import build_llm_input_from_merged_text


# =========================================================
# COLUNAS
# =========================================================

CSV_COLUMNS = [
    "N Interno", "Status", "Data Disputa", "Hora Disputa", "Órgão", "Tipo Licitação",
    "Resumo Switches", "Critério", "Local", "UASG", "Nº Pregão", "Cidade", "UF",
    "Valor Total Switches", "Vlr Total Edital", "Intervalo", "Exclusividade ME/EPP",
    "Endereço", "CEP", "Risco Identificado", "Habilitação jurídica",
    "HABILITAÇÃO FISCAL, SOCIAL E TRABALHISTA", "QUALIFICAÇÃO ECONÔMICO-FINANCEIRA",
    "QUALIFICAÇÃO TÉCNICA",
    "Item 1", "Preço item 1", "Quantidade item 1",
    "Item 2", "Preço item 2", "Quantidade item 2",
    "Item 3", "Preço item 3", "Quantidade item 3",
    "Item 4", "Preço item 4", "Quantidade item 4",
    "Item 5", "Preço item 5", "Quantidade item 5",
    "Item 6", "Preço item 6", "Quantidade item 6",
    "Item 7", "Preço item 7", "Quantidade item 7",
    "Item 8", "Preço item 8", "Quantidade item 8",
    "Item 9", "Preço item 9", "Quantidade item 9",
    "Item 10", "Preço item 10", "Quantidade item 10",
]

PLANILHA_COLUMNS = [
    "Selecionado", "URL DRIVE", "N interno", "Status", "Data Abertura", "Hora Abertura",
    "Órgão", "Tipo Licitação", "Resumo Switches", "Critério", "Portal", "UASG",
    "Nº Pregão", "Cidade", "UF", "Valor Total Switches", "Vlr Total Edital", "Intervalo",
    "Exclusividade ME/EPP", "Endereço", "CEP", "Risco Identificado", "Habilitação jurídica",
    "HABILITAÇÃO FISCAL, SOCIAL E TRABALHISTA", "QUALIFICAÇÃO ECONÔMICO-FINANCEIRA",
    "QUALIFICAÇÃO TÉCNICA",
    "Item 1", "Preço item 1", "Quantidade item 1",
    "Item 2", "Preço item 2", "Quantidade item 2",
    "Item 3", "Preço item 3", "Quantidade item 3",
    "Item 4", "Preço item 4", "Quantidade item 4",
    "Item 5", "Preço item 5", "Quantidade item 5",
    "Item 6", "Preço item 6", "Quantidade item 6",
    "Item 7", "Preço item 7", "Quantidade item 7",
    "Item 8", "Preço item 8", "Quantidade item 8",
    "Item 9", "Preço item 9", "Quantidade item 9",
    "Item 10", "Preço item 10", "Quantidade item 10",
]

BASE_RESULT_FIELDS = CSV_COLUMNS[:24]


# =========================================================
# SCHEMA OLLAMA
# =========================================================

OLLAMA_OUTPUT_SCHEMA: dict[str, Any] = {
    "type": "object",
    "additionalProperties": True,
    "properties": {
        field: {"type": "string"} for field in BASE_RESULT_FIELDS
    } | {
        "itens": {
            "type": "array",
            "maxItems": 10,
            "items": {
                "type": "object",
                "additionalProperties": True,
                "properties": {
                    "descricao": {"type": "string"},
                    "preco": {"type": "string"},
                    "quantidade": {"type": "string"},
                },
            },
        }
    },
}

SYSTEM_PROMPT = """
Você é um analisador técnico de editais de TI para revenda ME/EPP.

OBJETIVO
Analisar as evidências extraídas de todos os PDFs de um único edital e identificar potencial comercial para:
- transceiver e módulos ópticos autônomos
- switch 48 portas
- switch 24 portas
- switch 16 portas
- switch 8 portas

REGRAS CRÍTICAS
- Considere todos os arquivos como parte do mesmo edital.
- Não conclua ausência de switch só porque a palavra switch não aparece.
- Classifique como switch qualquer item com múltiplas portas ethernet, uplinks SFP, VLAN, PoE, gerenciamento, layer 2, layer 3 ou capacidade de comutação.
- Classifique transceiver como item autônomo somente quando houver item próprio, linha própria, quantidade própria, preço próprio, lote próprio ou CATMAT próprio.
- Se o transceiver aparecer apenas como incluso, acompanhado, embarcado, obrigatório, bundle, kit ou componente de switch/equipamento, NÃO liste como item separado.
- Nunca inferir quantidade de transceivers pelo número de portas do switch.
- Se houver switch e transceiver autônomo no mesmo edital, preserve ambos.

STATUS
- 🟢 se houver transceiver autônomo OU switch 48 portas OU switch 24 portas
- 🟡 se houver apenas switch 16 portas ou switch 8 portas
- 🔴 se não houver switch elegível nem transceiver autônomo

RISCO
Preencher "🚨 SERVIÇO/RISCO DETECTADO - [motivo curto]" quando houver instalação física, montagem em rack, configuração obrigatória, treinamento presencial, manutenção on-site, garantia on-site, SLA com deslocamento, visita técnica, entrega pulverizada, integração com rede existente, suporte presencial ou técnico certificado exigido.
Se não houver risco, preencher "Nenhum".

PREENCHIMENTO
- Use linguagem curta e objetiva.
- Se faltar dado, use "N/C".
- Não use ponto e vírgula dentro dos valores.
- Preencha até 10 itens de interesse.
- Prioridade: Transceiver autônomo > switch 48p > switch 24p > switch 16p > switch 8p.
- Se não houver item elegível, preencha itens com N/C.
- Resumo Switches deve listar switches e transceivers autônomos. Ex: "2x 48p -- 6x SFP+ 10Gb".
- Valor Total Switches = soma de preço unitário x quantidade dos itens elegíveis. Se não der para calcular, use N/C.

FORMATO DE SAÍDA
Retorne apenas JSON válido.
Sem markdown.
Sem comentários.
"""

PROMPT_TEMPLATE = """
Analise as evidências consolidadas abaixo. Elas foram extraídas dos PDFs do mesmo edital.

EVIDÊNCIAS DO EDITAL
{conteudo}
"""


# =========================================================
# PASTAS DE SAÍDA
# =========================================================

def _docling_output_dir() -> Path:
    return Path(__file__).resolve().parent / "docling"


def _json_output_dir() -> Path:
    return Path(__file__).resolve().parent / "json"


def _planilha_default_path() -> Path:
    return Path(__file__).resolve().parent / "planilha" / "planilha_editais.xlsx"


# =========================================================
# UTILITÁRIOS
# =========================================================

def sanitize_csv_value(value: Any) -> str:
    if value is None:
        return "N/C"
    if isinstance(value, (dict, list, tuple, set)):
        return "N/C"
    text = str(value).strip()
    if not text:
        return "N/C"
    text = text.replace(";", ",")
    text = " ".join(text.split())
    return text or "N/C"


def _safe_stem(name: str) -> str:
    stem = Path(name).stem
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._-")
    return stem or "document"


def _extract_n_interno_from_folder(folder_path: str | Path) -> str | None:
    name = Path(folder_path).name
    m = re.search(r"(\d{4}_\d{2}_\d{2}_\d+)", name)
    return m.group(1) if m else None


def _norm_key(key: str) -> str:
    k = unicodedata.normalize("NFKD", str(key))
    k = "".join(ch for ch in k if not unicodedata.combining(ch))
    k = k.lower().strip().replace("º", "o").replace("°", "o")
    k = re.sub(r"[^a-z0-9]+", " ", k)
    return " ".join(k.split())


def _is_nc(value: Any) -> bool:
    return sanitize_csv_value(value) == "N/C"


def _to_float_brl(value: str) -> float | None:
    text = sanitize_csv_value(value)
    if text == "N/C":
        return None
    text = text.replace("R$", "").strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except Exception:
        return None


def _format_brl(value: float) -> str:
    s = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


# =========================================================
# PDF / DOCLING
# =========================================================

def parse_edital_pdf(source: str | Path | bytes, filename: str | None = None) -> ParsedDocument:
    if filename is None:
        filename = Path(source).name if isinstance(source, (str, Path)) else "document.pdf"
    return parse_pdf(source, filename=filename)


def _parse_pdf_worker(pdf_path: str, result_queue: Any) -> None:
    try:
        parsed = parse_edital_pdf(pdf_path)
        result_queue.put({"ok": True, "text": parsed.full_text or ""})
    except Exception as exc:
        result_queue.put({"ok": False, "error": str(exc)})


def _parse_pdf_text_isolated(pdf_file: Path, timeout_sec: int = 600) -> str:
    ctx = mp.get_context("spawn")
    queue = ctx.Queue(maxsize=1)
    proc = ctx.Process(target=_parse_pdf_worker, args=(str(pdf_file), queue), daemon=True)
    proc.start()
    proc.join(timeout_sec)

    if proc.is_alive():
        proc.terminate()
        proc.join()
        raise TimeoutError(f"Timeout no parse de '{pdf_file.name}' após {timeout_sec}s")

    if proc.exitcode != 0:
        raise RuntimeError(f"Falha nativa no parse de '{pdf_file.name}' exitcode={proc.exitcode}")

    if queue.empty():
        raise RuntimeError(f"Parse de '{pdf_file.name}' finalizou sem retorno")

    payload = queue.get()
    if not isinstance(payload, dict) or not payload.get("ok"):
        raise RuntimeError(str(payload.get("error", "erro desconhecido")))

    return str(payload.get("text", ""))


def collect_pdf_files(folder_path: str | Path) -> list[Path]:
    folder = Path(folder_path)
    if not folder.exists():
        raise FileNotFoundError(f"Pasta não encontrada: {folder}")
    if not folder.is_dir():
        raise NotADirectoryError(f"O caminho informado não é uma pasta: {folder}")
    pdfs = sorted(p for p in folder.iterdir() if p.is_file() and p.suffix.lower() == ".pdf")
    if not pdfs:
        raise ValueError(f"Nenhum PDF encontrado na pasta: {folder}")
    return pdfs


def collect_pdf_sources(path_or_file: str | Path) -> list[Path]:
    p = Path(path_or_file)
    if not p.exists():
        raise FileNotFoundError(f"Caminho não encontrado: {p}")
    if p.is_file():
        if p.suffix.lower() != ".pdf":
            raise ValueError(f"Arquivo não é PDF: {p}")
        return [p]
    if p.is_dir():
        return collect_pdf_files(p)
    raise ValueError(f"Caminho inválido: {p}")


def _folder_has_pdf_files(folder_path: str | Path) -> bool:
    folder = Path(folder_path)
    return folder.exists() and folder.is_dir() and any(p.is_file() and p.suffix.lower() == ".pdf" for p in folder.iterdir())


def _is_pdf_file(path: str | Path) -> bool:
    p = Path(path)
    return p.exists() and p.is_file() and p.suffix.lower() == ".pdf"


def expand_input_folders(paths: list[str]) -> list[str]:
    expanded: list[str] = []
    seen: set[str] = set()
    for raw in paths:
        p = Path(raw)
        if _is_pdf_file(p):
            key = str(p.resolve())
            if key not in seen:
                seen.add(key)
                expanded.append(str(p))
            continue
        if _folder_has_pdf_files(p):
            key = str(p.resolve())
            if key not in seen:
                seen.add(key)
                expanded.append(str(p))
            continue
        if p.exists() and p.is_dir():
            matched = [d for d in sorted(p.iterdir()) if d.is_dir() and _folder_has_pdf_files(d)]
            if matched:
                for d in matched:
                    key = str(d.resolve())
                    if key not in seen:
                        seen.add(key)
                        expanded.append(str(d))
                continue
        key = str(p.resolve()) if p.exists() else str(p)
        if key not in seen:
            seen.add(key)
            expanded.append(str(p))
    return expanded


def _save_docling_markdown(pdf_file: Path, text: str) -> None:
    text = (text or "").strip()
    if not text:
        return
    out_dir = _docling_output_dir()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_name = f"{_safe_stem(pdf_file.parent.name)}__{_safe_stem(pdf_file.name)}.md"
    (out_dir / out_name).write_text(text, encoding="utf-8")


def parse_folder_as_single_edital(folder_path: str | Path) -> str:
    pdf_files = collect_pdf_sources(folder_path)
    docs_text: list[str] = []
    failed_files: list[tuple[str, str]] = []

    for pdf_file in pdf_files:
        try:
            text = _parse_pdf_text_isolated(pdf_file).strip()
            _save_docling_markdown(pdf_file, text)
            if text:
                docs_text.append(f"\n### ARQUIVO: {pdf_file.name}\n{text}")
        except Exception as exc:
            failed_files.append((pdf_file.name, str(exc)))
            print(f"Aviso: falha ao extrair '{pdf_file.name}': {exc}", file=sys.stderr)

    merged_text = "\n".join(docs_text).strip()
    if not merged_text:
        details = "; ".join(f"{n}: {e}" for n, e in failed_files[:3])
        raise ValueError(f"Nenhum texto foi extraído dos PDFs. {details}")
    return merged_text


# =========================================================
# JSON / NORMALIZAÇÃO
# =========================================================

def extract_first_json(text: str) -> dict[str, Any]:
    text = (text or "").strip()
    if text.startswith("```"):
        text = re.sub(r"^```[a-zA-Z0-9_-]*\s*", "", text)
        text = re.sub(r"\s*```\s*$", "", text).strip()
    try:
        data = json.loads(text)
        if isinstance(data, dict):
            return data
    except Exception:
        pass

    decoder = json.JSONDecoder()
    for start in sorted(set(m.start() for m in re.finditer(r"[\{\[]", text))):
        try:
            obj, _end = decoder.raw_decode(text[start:])
            if isinstance(obj, dict):
                return obj
        except Exception:
            continue
    raise ValueError("Não foi possível extrair JSON válido da resposta do modelo")


def ensure_result_shape(data: dict[str, Any]) -> dict[str, Any]:
    key_map = {_norm_key(k): k for k in data.keys()}
    synonyms: dict[str, list[str]] = {
        "N Interno": ["n interno", "numero interno", "no interno"],
        "Data Disputa": ["data disputa", "data abertura", "data da disputa"],
        "Hora Disputa": ["hora disputa", "hora abertura", "horario"],
        "Órgão": ["orgao", "entidade", "unidade"],
        "Tipo Licitação": ["tipo licitacao", "modalidade", "tipo"],
        "Resumo Switches": ["resumo switches", "resumo", "switches"],
        "Critério": ["criterio", "julgamento", "criterio de julgamento"],
        "Local": ["local", "portal", "site", "plataforma"],
        "Nº Pregão": ["no pregao", "n pregao", "numero pregao", "pregao"],
        "Cidade": ["cidade", "municipio"],
        "UF": ["uf", "estado"],
        "Valor Total Switches": ["valor total switches", "valor switches", "total switches"],
        "Vlr Total Edital": ["vlr total edital", "valor total edital", "valor edital", "total edital"],
        "Exclusividade ME/EPP": ["exclusividade me epp", "me epp", "me/epp", "exclusividade"],
        "Endereço": ["endereco", "endereco completo"],
        "Risco Identificado": ["risco identificado", "risco", "riscos"],
        "Habilitação jurídica": ["habilitacao juridica", "habilitacao"],
        "HABILITAÇÃO FISCAL, SOCIAL E TRABALHISTA": ["habilitacao fiscal social e trabalhista", "regularidade fiscal", "habilitacao fiscal"],
        "QUALIFICAÇÃO ECONÔMICO-FINANCEIRA": ["qualificacao economico financeira", "qualificacao economico-financeira"],
        "QUALIFICAÇÃO TÉCNICA": ["qualificacao tecnica"],
    }

    def get_field(field: str) -> Any:
        if field in data:
            return data.get(field)
        nk = _norm_key(field)
        if nk in key_map:
            return data.get(key_map[nk])
        for s in synonyms.get(field, []):
            ns = _norm_key(s)
            if ns in key_map:
                return data.get(key_map[ns])
        return "N/C"

    normalized: dict[str, Any] = {field: sanitize_csv_value(get_field(field)) for field in BASE_RESULT_FIELDS}

    itens = data.get("itens", data.get("items", data.get("Itens", [])))
    if not isinstance(itens, list):
        itens = []
    clean_items: list[dict[str, str]] = []
    for item in itens[:10]:
        if isinstance(item, dict):
            clean_items.append({
                "descricao": sanitize_csv_value(item.get("descricao", item.get("descrição", item.get("descricao_item", "N/C")))),
                "preco": sanitize_csv_value(item.get("preco", item.get("preço", item.get("valor", "N/C")))),
                "quantidade": sanitize_csv_value(item.get("quantidade", item.get("qtd", item.get("quant", "N/C")))),
            })
    normalized["itens"] = clean_items
    return normalized


# =========================================================
# FALLBACK / VALIDAÇÃO
# =========================================================

def _apply_rule_based_fallback(result: dict[str, Any], merged_text: str) -> dict[str, Any]:
    out = dict(result)
    low = merged_text.lower()
    itens = out.get("itens") if isinstance(out.get("itens"), list) else []

    if _is_nc(out.get("Status")):
        all_desc = " ".join(sanitize_csv_value(i.get("descricao", "")) for i in itens if isinstance(i, dict)).lower()
        if any(x in all_desc for x in ["48 portas", "48p", "24 portas", "24p", "transceiver", "transceptor", "sfp", "gbic"]):
            out["Status"] = "🟢"
        elif any(x in all_desc for x in ["16 portas", "16p", "8 portas", "8p"]):
            out["Status"] = "🟡"
        else:
            out["Status"] = "🔴"

    if _is_nc(out.get("Resumo Switches")) and itens:
        parts: list[str] = []
        total = 0.0
        total_ok = True
        for item in itens:
            if not isinstance(item, dict):
                continue
            desc = sanitize_csv_value(item.get("descricao", "N/C"))
            qty_s = sanitize_csv_value(item.get("quantidade", "N/C"))
            price_s = sanitize_csv_value(item.get("preco", "N/C"))
            m = re.search(r"\b(48|24|16|8)\s*(?:portas|p)\b", desc, flags=re.IGNORECASE)
            if m and qty_s != "N/C":
                parts.append(f"{qty_s}x {m.group(1)}p")
            elif re.search(r"\b(transceiver|transceptor|sfp\+?|gbic|qsfp)\b", desc, flags=re.IGNORECASE) and qty_s != "N/C":
                parts.append(f"{qty_s}x {desc}")
            try:
                qty = float(qty_s.replace(".", "").replace(",", "."))
                price = _to_float_brl(price_s)
                if price is None:
                    total_ok = False
                else:
                    total += qty * price
            except Exception:
                total_ok = False
        if parts:
            out["Resumo Switches"] = " -- ".join(parts[:10])
        if _is_nc(out.get("Valor Total Switches")) and total_ok and total > 0:
            out["Valor Total Switches"] = _format_brl(total)

    hora = sanitize_csv_value(out.get("Hora Disputa"))
    if hora != "N/C":
        m = re.match(r"^\s*(\d{1,2})[Hh](\d{2})?(?:min|m)?\s*$", hora)
        if m:
            out["Hora Disputa"] = f"{int(m.group(1)):02d}:{int(m.group(2) or 0):02d}"

    tipo = sanitize_csv_value(out.get("Tipo Licitação"))
    if tipo != "N/C" and "pregão eletrônico" in tipo.lower() and "registro de preços" in low and "srp" not in tipo.lower():
        out["Tipo Licitação"] = f"{tipo} SRP"

    risco = sanitize_csv_value(out.get("Risco Identificado"))
    if risco == "N/C":
        out["Risco Identificado"] = "Nenhum"

    return out


# =========================================================
# OLLAMA
# =========================================================

def call_ollama(
    prompt: str,
    model: str = "llama3.1:8b",
    host: str = "http://localhost:11434",
    timeout: int = 600,
    num_ctx: int = 8192,
    num_predict: int = 4096,
    retries: int = 2,
) -> str:
    url = f"{host.rstrip('/')}/api/generate"
    payload = {
        "model": model,
        "prompt": f"{SYSTEM_PROMPT}\n\n{prompt}",
        "stream": False,
        "format": OLLAMA_OUTPUT_SCHEMA,
        "options": {
            "temperature": 0.0,
            "num_ctx": int(num_ctx),
            "num_predict": int(num_predict),
        },
    }

    attempts = max(1, int(retries) + 1)
    transient_http = {429, 500, 502, 503, 504}
    last_exc: Exception | None = None

    for attempt in range(attempts):
        try:
            response = requests.post(url, json=payload, timeout=timeout)
        except requests.RequestException as exc:
            last_exc = exc
            if attempt < attempts - 1:
                time.sleep(min(2 ** attempt, 10))
                continue
            raise RuntimeError(f"Falha ao conectar no Ollama em {url}: {exc}") from exc

        if response.status_code >= 400:
            body = (response.text or "").strip()[:2000]
            if response.status_code in transient_http and attempt < attempts - 1:
                time.sleep(min(2 ** attempt, 10))
                continue
            raise RuntimeError(f"Ollama retornou HTTP {response.status_code}. Corpo: {body or 'N/C'}")

        try:
            data = response.json()
            return str(data.get("response", "")).strip()
        except Exception as exc:
            raise RuntimeError(f"Resposta do Ollama não é JSON válido: {(response.text or '')[:1000]}") from exc

    raise RuntimeError(f"Falha ao chamar Ollama: {last_exc}")


_SYSTEM_PROMPT_TOKENS = 1000
_RESPONSE_RESERVE_TOKENS = 1500
_CHARS_PER_TOKEN = 4


def _truncate_for_context(text: str, num_ctx: int) -> str:
    available = max(num_ctx - _SYSTEM_PROMPT_TOKENS - _RESPONSE_RESERVE_TOKENS, 1000)
    max_chars = max(available * _CHARS_PER_TOKEN, 4000)
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n\n[... TEXTO CORTADO POR LIMITE DE CONTEXTO ...]"


def analyze_edital_with_ollama(
    merged_text: str,
    model: str = "llama3.1:8b",
    host: str = "http://localhost:11434",
    num_ctx: int = 8192,
    timeout: int = 600,
    num_predict: int = 4096,
    retries: int = 2,
) -> dict[str, Any]:
    evidence_text = build_llm_input_from_merged_text(merged_text)
    safe_text = _truncate_for_context(evidence_text, num_ctx)
    prompt = PROMPT_TEMPLATE.format(conteudo=safe_text)

    raw_response = call_ollama(
        prompt=prompt,
        model=model,
        host=host,
        num_ctx=num_ctx,
        timeout=timeout,
        num_predict=num_predict,
        retries=retries,
    )
    parsed = extract_first_json(raw_response)
    normalized = ensure_result_shape(parsed)
    return _apply_rule_based_fallback(normalized, merged_text)


# =========================================================
# CSV / PLANILHA
# =========================================================

def normalize_result_to_csv_row(result: dict[str, Any]) -> dict[str, str]:
    row = {col: "N/C" for col in CSV_COLUMNS}
    for col in BASE_RESULT_FIELDS:
        row[col] = sanitize_csv_value(result.get(col, "N/C"))
    itens = result.get("itens", [])
    if not isinstance(itens, list):
        itens = []
    for idx in range(10):
        if idx < len(itens) and isinstance(itens[idx], dict):
            row[f"Item {idx + 1}"] = sanitize_csv_value(itens[idx].get("descricao", "N/C"))
            row[f"Preço item {idx + 1}"] = sanitize_csv_value(itens[idx].get("preco", "N/C"))
            row[f"Quantidade item {idx + 1}"] = sanitize_csv_value(itens[idx].get("quantidade", "N/C"))
    return row


def row_to_csv_line(row: dict[str, str]) -> str:
    return ";".join(sanitize_csv_value(row.get(col, "N/C")) for col in CSV_COLUMNS)


def build_csv_output(row: dict[str, str]) -> str:
    return ";".join(CSV_COLUMNS) + "\n" + row_to_csv_line(row)


def normalize_result_to_planilha_row(pipeline_row: dict[str, str], selecionado: str = "sim", url_drive: str = "") -> dict[str, str]:
    row = {col: "N/C" for col in PLANILHA_COLUMNS}
    row["Selecionado"] = sanitize_csv_value(selecionado)
    row["URL DRIVE"] = sanitize_csv_value(url_drive) if url_drive else ""
    mapping = {
        "N interno": "N Interno", "Status": "Status", "Data Abertura": "Data Disputa",
        "Hora Abertura": "Hora Disputa", "Órgão": "Órgão", "Tipo Licitação": "Tipo Licitação",
        "Resumo Switches": "Resumo Switches", "Critério": "Critério", "Portal": "Local",
        "UASG": "UASG", "Nº Pregão": "Nº Pregão", "Cidade": "Cidade", "UF": "UF",
        "Valor Total Switches": "Valor Total Switches", "Vlr Total Edital": "Vlr Total Edital",
        "Intervalo": "Intervalo", "Exclusividade ME/EPP": "Exclusividade ME/EPP",
        "Endereço": "Endereço", "CEP": "CEP", "Risco Identificado": "Risco Identificado",
        "Habilitação jurídica": "Habilitação jurídica",
        "HABILITAÇÃO FISCAL, SOCIAL E TRABALHISTA": "HABILITAÇÃO FISCAL, SOCIAL E TRABALHISTA",
        "QUALIFICAÇÃO ECONÔMICO-FINANCEIRA": "QUALIFICAÇÃO ECONÔMICO-FINANCEIRA",
        "QUALIFICAÇÃO TÉCNICA": "QUALIFICAÇÃO TÉCNICA",
    }
    for dst, src in mapping.items():
        row[dst] = sanitize_csv_value(pipeline_row.get(src, "N/C"))
    for i in range(1, 11):
        row[f"Item {i}"] = sanitize_csv_value(pipeline_row.get(f"Item {i}", "N/C"))
        row[f"Preço item {i}"] = sanitize_csv_value(pipeline_row.get(f"Preço item {i}", "N/C"))
        row[f"Quantidade item {i}"] = sanitize_csv_value(pipeline_row.get(f"Quantidade item {i}", "N/C"))
    return row


def planilha_row_to_csv_line(row: dict[str, str]) -> str:
    return ";".join((str(row.get(col, "")) if col == "URL DRIVE" else sanitize_csv_value(row.get(col, "N/C"))) for col in PLANILHA_COLUMNS)


def build_planilha_csv_output(row: dict[str, str]) -> str:
    return ";".join(PLANILHA_COLUMNS) + "\n" + planilha_row_to_csv_line(row)


def append_row_to_xlsx(planilha_path: str | Path, row: dict[str, str]) -> Path:
    planilha_path = Path(planilha_path)
    planilha_path.parent.mkdir(parents=True, exist_ok=True)
    if planilha_path.exists():
        wb = load_workbook(planilha_path)
        ws = wb.active
        if ws.max_row < 1:
            ws.append(PLANILHA_COLUMNS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Planilha"
        ws.append(PLANILHA_COLUMNS)
    ws.append([row.get(col, "") for col in PLANILHA_COLUMNS])
    wb.save(planilha_path)
    return planilha_path


def _save_analysis_json(folder_path: str | Path, result: dict[str, Any], model: str, host: str, num_ctx: int) -> Path:
    out_dir = _json_output_dir()
    out_dir.mkdir(parents=True, exist_ok=True)
    folder_tag = _safe_stem(Path(folder_path).name)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"{folder_tag}__{ts}.json"
    payload = {"model": model, "host": host, "num_ctx": num_ctx, "result": result}
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out_path


def build_error_pipeline_row(folder_path: str | Path, error_message: str) -> dict[str, str]:
    row = {col: "N/C" for col in CSV_COLUMNS}
    inferred_n = _extract_n_interno_from_folder(folder_path)
    if inferred_n:
        row["N Interno"] = inferred_n
    row["Status"] = "erro"
    row["Risco Identificado"] = sanitize_csv_value(f"Falha na análise: {error_message}")
    return row


# =========================================================
# PIPELINE
# =========================================================

def run_row(
    folder_path: str | Path,
    model: str,
    host: str,
    num_ctx: int,
    *,
    timeout: int,
    num_predict: int,
    retries: int,
) -> dict[str, str]:
    merged_text = parse_folder_as_single_edital(folder_path)
    release_converter()
    result = analyze_edital_with_ollama(
        merged_text=merged_text,
        model=model,
        host=host,
        num_ctx=num_ctx,
        timeout=timeout,
        num_predict=num_predict,
        retries=retries,
    )
    inferred_n = _extract_n_interno_from_folder(folder_path)
    if inferred_n:
        result["N Interno"] = inferred_n
    try:
        _save_analysis_json(folder_path, result, model, host, num_ctx)
    except Exception as exc:
        print(f"Aviso: falha ao salvar JSON: {exc}", file=sys.stderr)
    return normalize_result_to_csv_row(result)


def run(
    folder_path: str | Path,
    model: str,
    host: str,
    num_ctx: int,
    *,
    timeout: int,
    num_predict: int,
    retries: int,
) -> str:
    row = run_row(folder_path, model, host, num_ctx, timeout=timeout, num_predict=num_predict, retries=retries)
    return build_csv_output(row)


# =========================================================
# CLI
# =========================================================

def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

    parser = argparse.ArgumentParser(description="Analisa PDFs de editais usando Docling + extrator de evidências + Ollama.")
    parser.add_argument("pastas", nargs="+", type=str, help="Uma ou mais pastas contendo PDFs do edital.")
    parser.add_argument("--model", type=str, default="llama3.1:8b", help="Modelo do Ollama.")
    parser.add_argument("--host", type=str, default="http://localhost:11434", help="Host do Ollama.")
    parser.add_argument("--ctx", type=int, default=8192, help="Contexto num_ctx do Ollama.")
    parser.add_argument("--ollama-timeout", type=int, default=600, help="Timeout da chamada ao Ollama.")
    parser.add_argument("--ollama-retries", type=int, default=2, help="Retries para falhas temporárias.")
    parser.add_argument("--ollama-num-predict", type=int, default=4096, help="Limite de tokens da resposta.")
    parser.add_argument("--no-planilha", action="store_true", help="Não salva XLSX.")
    parser.add_argument("--planilha-arquivo", type=str, default=str(_planilha_default_path()), help="Caminho do XLSX.")
    parser.add_argument("--selecionado", type=str, default="sim", help="Valor da coluna Selecionado.")
    parser.add_argument("--url-drive", type=str, default="", help="Valor da coluna URL DRIVE.")
    parser.add_argument("--stdout-format", choices=["planilha", "pipeline"], default="planilha", help="Formato impresso no terminal.")
    parser.add_argument("--docling-low-memory", action="store_true", help="Ativa modo low-memory do Docling.")
    parser.add_argument("--docling-no-table", action="store_true", help="Desliga estrutura de tabelas do Docling.")
    parser.add_argument("--docling-no-ocr", action="store_true", help="Desliga OCR do Docling.")

    args = parser.parse_args()

    if args.docling_low_memory:
        os.environ["DOCLING_LOW_MEMORY"] = "1"
        os.environ.setdefault("DOCLING_PYPDF_FIRST", "1")
    if args.docling_no_table:
        os.environ["DOCLING_DO_TABLE_STRUCTURE"] = "0"
    if args.docling_no_ocr:
        os.environ["DOCLING_DO_OCR"] = "0"

    try:
        pastas = expand_input_folders(list(args.pastas))
        multi = len(pastas) > 1

        if multi:
            print(";".join(PLANILHA_COLUMNS if args.stdout_format == "planilha" else CSV_COLUMNS))

        last_saved: Path | None = None
        for pasta in pastas:
            try:
                row = run_row(
                    folder_path=pasta,
                    model=args.model,
                    host=args.host,
                    num_ctx=args.ctx,
                    timeout=args.ollama_timeout,
                    num_predict=args.ollama_num_predict,
                    retries=args.ollama_retries,
                )
            except Exception as exc:
                row = build_error_pipeline_row(pasta, str(exc))
                print(f"Aviso: falha em '{pasta}': {exc}", file=sys.stderr)

            planilha_row = normalize_result_to_planilha_row(row, selecionado=args.selecionado, url_drive=args.url_drive)

            if args.stdout_format == "planilha":
                line = planilha_row_to_csv_line(planilha_row)
                if multi:
                    print(line)
                else:
                    print(build_planilha_csv_output(planilha_row))
            else:
                line = row_to_csv_line(row)
                if multi:
                    print(line)
                else:
                    print(build_csv_output(row))

            if not args.no_planilha:
                last_saved = append_row_to_xlsx(args.planilha_arquivo, planilha_row)

        if last_saved is not None:
            print(f"\nPlanilha salva em: {last_saved}")

    except Exception as exc:
        print(f"Erro: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
