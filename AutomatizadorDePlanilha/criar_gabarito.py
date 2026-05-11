"""
Importa linhas de uma planilha .xlsx corrigida para o gabarito JSON
usado pelo pipeline com --resultado-identico.

Uso:
    python criar_gabarito.py                              # usa paths padrão
    python criar_gabarito.py --planilha minha.xlsx        # planilha específica
    python criar_gabarito.py --gabarito outro.json        # gabarito de saída
    python criar_gabarito.py --sobrescrever               # sobrescreve entradas existentes
    python criar_gabarito.py --limpar                     # gera gabarito novo do zero

A planilha deve ter cabeçalho com as colunas no formato PLANILHA_COLUMNS do pipeline.
A coluna "N interno" é obrigatória para identificar o edital.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


PLANILHA_COLUMNS = [
    "Selecionado",
    "URL DRIVE",
    "N interno",
    "Status",
    "Data Abertura",
    "Hora Abertura",
    "Órgão",
    "Tipo Licitação",
    "Resumo Switches",
    "Critério",
    "Portal",
    "UASG",
    "Nº Pregão",
    "Cidade",
    "UF",
    "Valor Total Switches",
    "Vlr Total Edital",
    "Intervalo",
    "Exclusividade ME/EPP",
    "Endereço",
    "CEP",
    "Risco Identificado",
    "Habilitação jurídica",
    "HABILITAÇÃO FISCAL, SOCIAL E TRABALHISTA",
    "QUALIFICAÇÃO ECONÔMICO-FINANCEIRA",
    "QUALIFICAÇÃO TÉCNICA",
    "Item 1",
    "Preço item 1",
    "Quantidade item 1",
    "Item 2",
    "Preço item 2",
    "Quantidade item 2",
    "Item 3",
    "Preço item 3",
    "Quantidade item 3",
    "Item 4",
    "Preço item 4",
    "Quantidade item 4",
    "Item 5",
    "Preço item 5",
    "Quantidade item 5",
    "Item 6",
    "Preço item 6",
    "Quantidade item 6",
    "Item 7",
    "Preço item 7",
    "Quantidade item 7",
    "Item 8",
    "Preço item 8",
    "Quantidade item 8",
    "Item 9",
    "Preço item 9",
    "Quantidade item 9",
    "Item 10",
    "Preço item 10",
    "Quantidade item 10",
]


def _cell_str(value) -> str:
    if value is None:
        return "N/C"
    text = str(value).strip()
    return text if text else "N/C"


def _default_planilha() -> Path:
    here = Path(__file__).resolve().parent
    # Tenta a planilha de resultado idêntico primeiro, depois a editais_unico
    for name in ("planilha_resultado_identico.xlsx", "planilha_editais_unico.xlsx", "planilha_editais.xlsx"):
        path = here / "planilha" / name
        if path.exists():
            return path
    return here / "planilha" / "planilha_resultado_identico.xlsx"


def _default_gabarito() -> Path:
    return Path(__file__).resolve().parent / "resultado_identico_gabarito.json"


def load_gabarito(path: Path) -> list[dict]:
    if not path.exists():
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def read_planilha_rows(planilha_path: Path) -> list[dict[str, str]]:
    wb = load_workbook(planilha_path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ValueError("Planilha vazia — nenhuma linha encontrada.")

    # Normaliza cabeçalho (strip + lowercase para encontrar colunas)
    header = [str(c).strip() if c is not None else "" for c in header_row]
    header_lower = [h.lower() for h in header]

    def _col_idx(name: str) -> int | None:
        try:
            return header_lower.index(name.lower())
        except ValueError:
            return None

    col_map: dict[str, int] = {}
    for col in PLANILHA_COLUMNS:
        idx = _col_idx(col)
        if idx is not None:
            col_map[col] = idx

    n_interno_idx = col_map.get("N interno")
    if n_interno_idx is None:
        raise ValueError("Coluna 'N interno' não encontrada na planilha.")

    result: list[dict[str, str]] = []
    for raw_row in rows_iter:
        raw_list = list(raw_row)
        n_int = _cell_str(raw_list[n_interno_idx] if n_interno_idx < len(raw_list) else None)
        if n_int == "N/C" or not n_int:
            continue  # pula linhas sem N interno

        row: dict[str, str] = {"N interno": n_int}
        for col, idx in col_map.items():
            if col == "N interno":
                continue
            val = raw_list[idx] if idx < len(raw_list) else None
            row[col] = _cell_str(val)

        # Garante que todas as colunas existam
        for col in PLANILHA_COLUMNS:
            if col not in row:
                row[col] = "N/C" if col != "URL DRIVE" else ""

        result.append(row)

    wb.close()
    return result


def merge_into_gabarito(
    existing: list[dict],
    new_rows: list[dict[str, str]],
    *,
    sobrescrever: bool,
) -> tuple[list[dict], int, int]:
    by_n: dict[str, int] = {}
    for i, entry in enumerate(existing):
        n = str(entry.get("N interno", "")).strip()
        if n:
            by_n[n] = i

    added = 0
    updated = 0
    result = list(existing)

    for row in new_rows:
        n = row.get("N interno", "").strip()
        if not n:
            continue

        if n in by_n:
            if sobrescrever:
                result[by_n[n]] = row
                updated += 1
        else:
            by_n[n] = len(result)
            result.append(row)
            added += 1

    return result, added, updated


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa planilha corrigida para o gabarito JSON.")
    parser.add_argument("--planilha", type=str, default=None, help="Caminho da planilha .xlsx")
    parser.add_argument("--gabarito", type=str, default=None, help="Caminho do gabarito JSON de saída")
    parser.add_argument("--sobrescrever", action="store_true", help="Sobrescreve entradas existentes no gabarito")
    parser.add_argument("--limpar", action="store_true", help="Ignora gabarito existente e gera do zero")
    args = parser.parse_args()

    planilha_path = Path(args.planilha) if args.planilha else _default_planilha()
    gabarito_path = Path(args.gabarito) if args.gabarito else _default_gabarito()

    if not planilha_path.exists():
        print(f"Erro: planilha não encontrada em '{planilha_path}'", file=sys.stderr)
        sys.exit(1)

    print(f"Lendo planilha: {planilha_path}")
    try:
        new_rows = read_planilha_rows(planilha_path)
    except Exception as exc:
        print(f"Erro ao ler planilha: {exc}", file=sys.stderr)
        sys.exit(1)

    print(f"  {len(new_rows)} linhas com N interno lidas.")

    existing = [] if args.limpar else load_gabarito(gabarito_path)
    print(f"Gabarito existente: {len(existing)} entradas.")

    merged, added, updated = merge_into_gabarito(
        existing, new_rows, sobrescrever=args.sobrescrever or args.limpar
    )

    gabarito_path.parent.mkdir(parents=True, exist_ok=True)
    gabarito_path.write_text(
        json.dumps(merged, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Gabarito salvo em: {gabarito_path}")
    print(f"  Adicionadas: {added} | Atualizadas: {updated} | Total: {len(merged)}")


if __name__ == "__main__":
    main()
