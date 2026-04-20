from __future__ import annotations

import os
import re
import sys
import json
import argparse
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook


# =========================================================
# AJUSTE DE IMPORT DO BACKEND
# =========================================================

def _ensure_backend_on_syspath() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    backend_dir = repo_root / "backend"
    backend_str = str(backend_dir)

    if backend_str not in sys.path:
        sys.path.insert(0, backend_str)


_ensure_backend_on_syspath()

from app.pipeline.docling_parser import ParsedDocument, parse_pdf  # type: ignore


# =========================================================
# CONFIG
# =========================================================

CSV_COLUMNS = [
    "N Interno",
    "Status",
    "Data Disputa",
    "Hora Disputa",
    "Órgão",
    "Tipo Licitação",
    "Resumo Switches",
    "Critério",
    "Local",
    "UASG",
    "Nº Pregão",
    "Cidade",
    "UF",
    "Valor Total Switches",
    "Vlr Total Edital",
    "Intervalo",
    "Exclusividade ME/EPP",
    "Endereço",
    "CEP",
    "Risco Identificado",
    "Habilitação jurídica",
    "HABILITAÇÃO FISCAL, SOCIAL E TRABALHISTA",
    "QUALIFICAÇÃO ECONÔMICO-FINANCEIRA",
    "QUALIFICAÇÃO TÉCNICA",
    "Item 1",
    "Preço item 1",
    "Quantidade item 1",
    "Item 2",
    "Preço item 2",
    "Quantidade item 2",
    "Item 3",
    "Preço item 3",
    "Quantidade item 3",
    "Item 4",
    "Preço item 4",
    "Quantidade item 4",
    "Item 5",
    "Preço item 5",
    "Quantidade item 5",
    "Item 6",
    "Preço item 6",
    "Quantidade item 6",
    "Item 7",
    "Preço item 7",
    "Quantidade item 7",
    "Item 8",
    "Preço item 8",
    "Quantidade item 8",
    "Item 9",
    "Preço item 9",
    "Quantidade item 9",
    "Item 10",
    "Preço item 10",
    "Quantidade item 10",
]

PLANILHA_COLUMNS = [
    "Selecionado",
    "URL DRIVE",
    "N interno",
    "Status",
    "Data Abertura",
    "Hora Abertura",
    "Órgão",
    "Tipo Licitação",
    "Resumo Switches",
    "Critério",
    "Portal",
    "UASG",
    "Nº Pregão",
    "Cidade",
    "UF",
    "Valor Total Switches",
    "Vlr Total Edital",
    "Intervalo",
    "Exclusividade ME/EPP",
    "Endereço",
    "CEP",
    "Risco Identificado",
    "Habilitação jurídica",
    "HABILITAÇÃO FISCAL, SOCIAL E TRABALHISTA",
    "QUALIFICAÇÃO ECONÔMICO-FINANCEIRA",
    "QUALIFICAÇÃO TÉCNICA",
    "Item 1",
    "Preço item 1",
    "Quantidade item 1",
    "Item 2",
    "Preço item 2",
    "Quantidade item 2",
    "Item 3",
    "Preço item 3",
    "Quantidade item 3",
    "Item 4",
    "Preço item 4",
    "Quantidade item 4",
    "Item 5",
    "Preço item 5",
    "Quantidade item 5",
    "Item 6",
    "Preço item 6",
    "Quantidade item 6",
    "Item 7",
    "Preço item 7",
    "Quantidade item 7",
    "Item 8",
    "Preço item 8",
    "Quantidade item 8",
    "Item 9",
    "Preço item 9",
    "Quantidade item 9",
    "Item 10",
    "Preço item 10",
    "Quantidade item 10",
]

BASE_RESULT_FIELDS = [
    "N Interno",
    "Status",
    "Data Disputa",
    "Hora Disputa",
    "Órgão",
    "Tipo Licitação",
    "Resumo Switches",
    "Critério",
    "Local",
    "UASG",
    "Nº Pregão",
    "Cidade",
    "UF",
    "Valor Total Switches",
    "Vlr Total Edital",
    "Intervalo",
    "Exclusividade ME/EPP",
    "Endereço",
    "CEP",
    "Risco Identificado",
    "Habilitação jurídica",
    "HABILITAÇÃO FISCAL, SOCIAL E TRABALHISTA",
    "QUALIFICAÇÃO ECONÔMICO-FINANCEIRA",
    "QUALIFICAÇÃO TÉCNICA",
]

SYSTEM_PROMPT = """
Você é um analisador técnico de editais de TI para revenda ME/EPP.

OBJETIVO
Analisar todos os PDFs recebidos como um único edital e identificar potencial comercial para:
- transceiver e módulos ópticos
- transceptor
- switch 48 portas
- switch 24 portas
- switch 16 portas
- switch 8 portas

REGRAS DE NEGÓCIO
- Sempre considerar todos os arquivos como parte do mesmo edital.
- Fazer varredura semântica completa em edital, termo de referência, relação de itens, anexos técnicos, tabelas e imagens extraídas.
- Não concluir ausência de switch apenas porque a palavra switch não aparece.
- Classificar como switch qualquer item com características compatíveis, como múltiplas portas ethernet, uplinks SFP, VLAN, PoE, gerenciamento, layer 2, layer 3 ou capacidade de comutação.
- Classificar como transceiver apenas quando houver item próprio, linha própria, quantitativo próprio ou preço próprio.
- Não classificar como item autônomo quando o transceiver estiver apenas incluído, embarcado, integrado, acompanhado, obrigatório para ativação de portas, bundle, kit, solução, lote ou composição de outro equipamento.
- Nunca inferir quantidade de transceivers a partir do número de portas do switch.
- Nunca criar item de transceiver sem evidência documental de autonomia comercial.
- Se houver switch e transceiver autônomo no mesmo edital, preservar ambos na saída.
- Nunca omitir transceiver autônomo por existir switch.

STATUS
- 🟢 se houver transceiver autônomo ou switch 48 portas ou switch 24 portas
- 🟡 se houver apenas switch 16 portas ou switch 8 portas
- 🔴 se não houver switch elegível nem transceiver autônomo

RISCO
Preencher:
- "🚨 SERVIÇO/RISCO DETECTADO - [motivo]"
quando houver instalação física, montagem em rack, configuração obrigatória, treinamento presencial, manutenção on-site, garantia on-site, SLA com deslocamento, visita técnica, entrega pulverizada, integração com rede existente, suporte presencial, técnico certificado exigido para execução ou exigência incompatível com revenda.
- Se não houver risco, preencher "Nenhum".

PREENCHIMENTO
- Usar linguagem curta e objetiva.
- Se faltar dado, usar "N/C".
- Não usar ponto e vírgula dentro dos valores.
- Preencher até 10 itens de interesse.
- Prioridade: Transceiver autônomo, switch 48p, switch 24p, switch 16p, switch 8p.
- Se não houver item elegível, preencher todos os itens com N/C.
- "Resumo Switches" deve incluir switches e transceivers autônomos.
- Não incluir transceivers embarcados ou apenas acessórios.
- "Valor Total Switches" deve somar valor unitário x quantidade dos itens elegíveis.
- Se não houver preço suficiente para calcular, usar N/C.

FORMATO DE SAÍDA
Retorne apenas JSON válido.
Sem markdown.
Sem comentários.
Sem texto antes ou depois.

ESTRUTURA OBRIGATÓRIA DO JSON
{
  "N Interno": "string",
  "Status": "string",
  "Data Disputa": "string",
  "Hora Disputa": "string",
  "Órgão": "string",
  "Tipo Licitação": "string",
  "Resumo Switches": "string",
  "Critério": "string",
  "Local": "string",
  "UASG": "string",
  "Nº Pregão": "string",
  "Cidade": "string",
  "UF": "string",
  "Valor Total Switches": "string",
  "Vlr Total Edital": "string",
  "Intervalo": "string",
  "Exclusividade ME/EPP": "string",
  "Endereço": "string",
  "CEP": "string",
  "Risco Identificado": "string",
  "Habilitação jurídica": "string",
  "HABILITAÇÃO FISCAL, SOCIAL E TRABALHISTA": "string",
  "QUALIFICAÇÃO ECONÔMICO-FINANCEIRA": "string",
  "QUALIFICAÇÃO TÉCNICA": "string",
  "itens": [
    {
      "descricao": "string",
      "preco": "string",
      "quantidade": "string"
    }
  ]
}
"""

PROMPT_TEMPLATE = """
Analise o edital consolidado abaixo. Todos os arquivos pertencem ao mesmo processo licitatório.

CONTEÚDO DO EDITAL
{conteudo}
"""


# =========================================================
# PARSER DOS PDFS
# =========================================================

def parse_edital_pdf(source: str | Path | bytes, filename: str | None = None) -> ParsedDocument:
    if filename is None:
        filename = Path(source).name if isinstance(source, (str, Path)) else "document.pdf"
    return parse_pdf(source, filename=filename)


def collect_pdf_files(folder_path: str | Path) -> list[Path]:
    folder = Path(folder_path)
    if not folder.exists():
        raise FileNotFoundError(f"Pasta não encontrada: {folder}")
    if not folder.is_dir():
        raise NotADirectoryError(f"O caminho informado não é uma pasta: {folder}")

    pdfs = sorted([p for p in folder.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"])
    if not pdfs:
        raise ValueError(f"Nenhum PDF encontrado na pasta: {folder}")

    return pdfs


def parse_folder_as_single_edital(folder_path: str | Path) -> str:
    pdf_files = collect_pdf_files(folder_path)
    docs_text: list[str] = []

    for pdf_file in pdf_files:
        parsed = parse_edital_pdf(pdf_file)
        text = (parsed.full_text or "").strip()

        if text:
            docs_text.append(f"\n### ARQUIVO: {pdf_file.name}\n{text}")

    merged_text = "\n".join(docs_text).strip()

    if not merged_text:
        raise ValueError("Nenhum texto foi extraído dos PDFs da pasta.")

    return merged_text


# =========================================================
# LIMPEZA / JSON
# =========================================================

def sanitize_csv_value(value: Any) -> str:
    if value is None:
        return "N/C"

    text = str(value).strip()
    if not text:
        return "N/C"

    text = text.replace(";", ",")
    text = " ".join(text.split())
    return text


def extract_first_json(text: str) -> dict[str, Any]:
    text = text.strip()

    # tentativa 1: json direto
    try:
        data = json.loads(text)
        if isinstance(data, dict):
            return data
    except Exception:
        pass

    # tentativa 2: encontrar primeiro bloco {...}
    match = re.search(r"\{.*\}", text, flags=re.DOTALL)
    if match:
        candidate = match.group(0)
        try:
            data = json.loads(candidate)
            if isinstance(data, dict):
                return data
        except Exception:
            pass

    import logging
    logging.getLogger(__name__).warning(
        "JSON inválido. Resposta bruta (primeiros 500 chars): %s", text[:500]
    )
    raise ValueError("Não foi possível extrair JSON válido da resposta do modelo.")


def ensure_result_shape(data: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}

    for field in BASE_RESULT_FIELDS:
        normalized[field] = sanitize_csv_value(data.get(field, "N/C"))

    itens = data.get("itens", [])
    if not isinstance(itens, list):
        itens = []

    clean_items: list[dict[str, str]] = []
    for item in itens[:10]:
        if not isinstance(item, dict):
            continue
        clean_items.append({
            "descricao": sanitize_csv_value(item.get("descricao", "N/C")),
            "preco": sanitize_csv_value(item.get("preco", "N/C")),
            "quantidade": sanitize_csv_value(item.get("quantidade", "N/C")),
        })

    normalized["itens"] = clean_items
    return normalized


# =========================================================
# OLLAMA
# =========================================================

def call_ollama(
    prompt: str,
    model: str = "llama3.1:8b",
    host: str = "http://localhost:11434",
    timeout: int = 600,
    num_ctx: int = 8192,
) -> str:
    url = f"{host.rstrip('/')}/api/generate"

    payload = {
        "model": model,
        "prompt": f"{SYSTEM_PROMPT}\n\n{prompt}",
        "stream": False,
        "options": {
            "temperature": 0.1,
            "num_ctx": int(num_ctx),
            "num_predict": 2048,
        },
        "format": "json",
    }

    try:
        response = requests.post(url, json=payload, timeout=timeout)
    except requests.RequestException as exc:
        raise RuntimeError(f"Falha ao conectar no Ollama em {url}: {exc}") from exc

    if response.status_code >= 400:
        body = (response.text or "").strip()
        if len(body) > 2000:
            body = body[:2000] + "..."
        raise RuntimeError(
            "Ollama retornou erro HTTP "
            f"{response.status_code} ao gerar resposta. "
            f"Corpo: {body or 'N/C'}"
        )

    try:
        data = response.json()
    except Exception as exc:
        snippet = (response.text or "").strip()
        if len(snippet) > 2000:
            snippet = snippet[:2000] + "..."
        raise RuntimeError(f"Resposta do Ollama não é JSON válido. Trecho: {snippet or 'N/C'}") from exc

    raw = str(data.get("response", "")).strip()
    if not raw:
        import logging
        logging.getLogger(__name__).warning(
            "Ollama retornou resposta vazia. done_reason=%s eval_count=%s",
            data.get("done_reason"), data.get("eval_count"),
        )
    return raw


_SYSTEM_PROMPT_TOKENS = 900   # estimativa conservadora
_RESPONSE_RESERVE_TOKENS = 1500  # espaço para JSON de saída completo (~10 itens)
_CHARS_PER_TOKEN = 4


def _truncate_for_context(text: str, num_ctx: int) -> str:
    available = num_ctx - _SYSTEM_PROMPT_TOKENS - _RESPONSE_RESERVE_TOKENS
    max_chars = max(available * _CHARS_PER_TOKEN, 4000)
    if len(text) <= max_chars:
        return text

    # Mantém o início (cabeçalho/identificação) e o fim (itens/anexos)
    half = max_chars // 2
    truncated = text[:half] + "\n\n[... TRECHO CENTRAL OMITIDO POR LIMITE DE CONTEXTO ...]\n\n" + text[-half:]
    import logging
    logging.getLogger(__name__).warning(
        "Texto truncado de %d para ~%d chars para caber em num_ctx=%d",
        len(text), max_chars, num_ctx,
    )
    return truncated


def analyze_edital_with_ollama(
    merged_text: str,
    model: str = "llama3.1:8b",
    host: str = "http://localhost:11434",
    num_ctx: int = 8192,
) -> dict[str, Any]:
    safe_text = _truncate_for_context(merged_text, num_ctx)
    prompt = PROMPT_TEMPLATE.format(conteudo=safe_text)
    raw_response = call_ollama(prompt=prompt, model=model, host=host, num_ctx=num_ctx)
    parsed = extract_first_json(raw_response)
    return ensure_result_shape(parsed)


# =========================================================
# FORMATADOR CSV
# =========================================================

def normalize_result_to_csv_row(result: dict[str, Any]) -> dict[str, str]:
    row: dict[str, str] = {col: "N/C" for col in CSV_COLUMNS}

    for col in BASE_RESULT_FIELDS:
        row[col] = sanitize_csv_value(result.get(col, "N/C"))

    itens = result.get("itens", [])
    if not isinstance(itens, list):
        itens = []

    for idx in range(10):
        desc_col = f"Item {idx + 1}"
        price_col = f"Preço item {idx + 1}"
        qty_col = f"Quantidade item {idx + 1}"

        if idx < len(itens):
            item = itens[idx]
            if isinstance(item, dict):
                row[desc_col] = sanitize_csv_value(item.get("descricao", "N/C"))
                row[price_col] = sanitize_csv_value(item.get("preco", "N/C"))
                row[qty_col] = sanitize_csv_value(item.get("quantidade", "N/C"))
            else:
                row[desc_col] = "N/C"
                row[price_col] = "N/C"
                row[qty_col] = "N/C"
        else:
            row[desc_col] = "N/C"
            row[price_col] = "N/C"
            row[qty_col] = "N/C"

    return row


def row_to_csv_line(row: dict[str, str]) -> str:
    return ";".join(sanitize_csv_value(row.get(col, "N/C")) for col in CSV_COLUMNS)


def build_csv_output(row: dict[str, str]) -> str:
    header = ";".join(CSV_COLUMNS)
    line = row_to_csv_line(row)
    return f"{header}\n{line}"


def _planilha_default_path() -> Path:
    here = Path(__file__).resolve().parent
    return here / "planilha" / "planilha_editais.xlsx"


def normalize_result_to_planilha_row(
    pipeline_row: dict[str, str],
    selecionado: str = "sim",
    url_drive: str = "",
) -> dict[str, str]:
    row: dict[str, str] = {col: "N/C" for col in PLANILHA_COLUMNS}
    row["Selecionado"] = sanitize_csv_value(selecionado)
    row["URL DRIVE"] = sanitize_csv_value(url_drive) if url_drive else ""

    row["N interno"] = sanitize_csv_value(pipeline_row.get("N Interno", "N/C"))
    row["Status"] = sanitize_csv_value(pipeline_row.get("Status", "N/C"))
    row["Data Abertura"] = sanitize_csv_value(pipeline_row.get("Data Disputa", "N/C"))
    row["Hora Abertura"] = sanitize_csv_value(pipeline_row.get("Hora Disputa", "N/C"))
    row["Órgão"] = sanitize_csv_value(pipeline_row.get("Órgão", "N/C"))
    row["Tipo Licitação"] = sanitize_csv_value(pipeline_row.get("Tipo Licitação", "N/C"))
    row["Resumo Switches"] = sanitize_csv_value(pipeline_row.get("Resumo Switches", "N/C"))
    row["Critério"] = sanitize_csv_value(pipeline_row.get("Critério", "N/C"))
    # A planilha tem "Portal"; no pipeline esse campo geralmente vem em "Local".
    row["Portal"] = sanitize_csv_value(pipeline_row.get("Local", "N/C"))
    row["UASG"] = sanitize_csv_value(pipeline_row.get("UASG", "N/C"))
    row["Nº Pregão"] = sanitize_csv_value(pipeline_row.get("Nº Pregão", "N/C"))
    row["Cidade"] = sanitize_csv_value(pipeline_row.get("Cidade", "N/C"))
    row["UF"] = sanitize_csv_value(pipeline_row.get("UF", "N/C"))
    row["Valor Total Switches"] = sanitize_csv_value(pipeline_row.get("Valor Total Switches", "N/C"))
    row["Vlr Total Edital"] = sanitize_csv_value(pipeline_row.get("Vlr Total Edital", "N/C"))
    row["Intervalo"] = sanitize_csv_value(pipeline_row.get("Intervalo", "N/C"))
    row["Exclusividade ME/EPP"] = sanitize_csv_value(pipeline_row.get("Exclusividade ME/EPP", "N/C"))
    row["Endereço"] = sanitize_csv_value(pipeline_row.get("Endereço", "N/C"))
    row["CEP"] = sanitize_csv_value(pipeline_row.get("CEP", "N/C"))
    row["Risco Identificado"] = sanitize_csv_value(pipeline_row.get("Risco Identificado", "N/C"))
    row["Habilitação jurídica"] = sanitize_csv_value(pipeline_row.get("Habilitação jurídica", "N/C"))
    row["HABILITAÇÃO FISCAL, SOCIAL E TRABALHISTA"] = sanitize_csv_value(
        pipeline_row.get("HABILITAÇÃO FISCAL, SOCIAL E TRABALHISTA", "N/C")
    )
    row["QUALIFICAÇÃO ECONÔMICO-FINANCEIRA"] = sanitize_csv_value(
        pipeline_row.get("QUALIFICAÇÃO ECONÔMICO-FINANCEIRA", "N/C")
    )
    row["QUALIFICAÇÃO TÉCNICA"] = sanitize_csv_value(pipeline_row.get("QUALIFICAÇÃO TÉCNICA", "N/C"))

    for i in range(1, 11):
        row[f"Item {i}"] = sanitize_csv_value(pipeline_row.get(f"Item {i}", "N/C"))
        row[f"Preço item {i}"] = sanitize_csv_value(pipeline_row.get(f"Preço item {i}", "N/C"))
        row[f"Quantidade item {i}"] = sanitize_csv_value(pipeline_row.get(f"Quantidade item {i}", "N/C"))

    return row


def append_row_to_xlsx(planilha_path: str | Path, row: dict[str, str]) -> Path:
    planilha_path = Path(planilha_path)
    planilha_path.parent.mkdir(parents=True, exist_ok=True)

    if planilha_path.exists():
        wb = load_workbook(planilha_path)
        ws = wb.active
        # Se a planilha existir mas estiver vazia, garante o cabeçalho
        if ws.max_row < 1:
            ws.append(PLANILHA_COLUMNS)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Planilha"
        ws.append(PLANILHA_COLUMNS)

    ws.append([row.get(col, "") for col in PLANILHA_COLUMNS])
    wb.save(planilha_path)
    return planilha_path


# =========================================================
# PIPELINE
# =========================================================

def run(folder_path: str | Path, model: str, host: str, num_ctx: int) -> str:
    merged_text = parse_folder_as_single_edital(folder_path)
    result = analyze_edital_with_ollama(merged_text=merged_text, model=model, host=host, num_ctx=num_ctx)
    row = normalize_result_to_csv_row(result)
    return build_csv_output(row)


def run_row(folder_path: str | Path, model: str, host: str, num_ctx: int) -> dict[str, str]:
    merged_text = parse_folder_as_single_edital(folder_path)
    result = analyze_edital_with_ollama(merged_text=merged_text, model=model, host=host, num_ctx=num_ctx)
    return normalize_result_to_csv_row(result)


# =========================================================
# CLI
# =========================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Analisa PDFs de uma pasta como um único edital usando Ollama."
    )
    parser.add_argument(
        "pasta",
        type=str,
        help="Caminho da pasta contendo os PDFs do edital.",
    )
    parser.add_argument(
        "--model",
        type=str,
        default="llama3.1:8b",
        help="Modelo do Ollama. Ex.: llama3.1:8b",
    )
    parser.add_argument(
        "--host",
        type=str,
        default="http://localhost:11434",
        help="Host do Ollama.",
    )
    parser.add_argument(
        "--ctx",
        type=int,
        default=8192,
        help="Contexto (num_ctx) para o Ollama. Valores muito altos podem causar erro 500 por falta de memória.",
    )
    parser.add_argument(
        "--no-planilha",
        action="store_true",
        help="Se informado, NÃO salva/atualiza a planilha .xlsx (padrão: salva automaticamente).",
    )
    parser.add_argument(
        "--planilha-arquivo",
        type=str,
        default=str(_planilha_default_path()),
        help="Caminho do arquivo .xlsx de saída (padrão: AnalisadorDeEditais/planilha/planilha_editais.xlsx).",
    )
    parser.add_argument(
        "--selecionado",
        type=str,
        default="sim",
        help="Valor da coluna 'Selecionado' na planilha (ex.: sim/nao).",
    )
    parser.add_argument(
        "--url-drive",
        type=str,
        default="",
        help="Valor da coluna 'URL DRIVE' na planilha.",
    )

    args = parser.parse_args()

    try:
        row = run_row(folder_path=args.pasta, model=args.model, host=args.host, num_ctx=args.ctx)
        csv_output = build_csv_output(row)
        print(csv_output)

        if not args.no_planilha:
            planilha_row = normalize_result_to_planilha_row(
                row,
                selecionado=args.selecionado,
                url_drive=args.url_drive,
            )
            saved_path = append_row_to_xlsx(args.planilha_arquivo, planilha_row)
            print(f"\nPlanilha salva em: {saved_path}")
    except Exception as exc:
        print(f"Erro: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()