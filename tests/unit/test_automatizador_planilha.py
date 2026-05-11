from __future__ import annotations

import importlib.util
import sys
import types
from pathlib import Path

from openpyxl import load_workbook


def _load_pipeline_module(monkeypatch):
    if "app" not in sys.modules:
        monkeypatch.setitem(sys.modules, "app", types.ModuleType("app"))
    if "app.pipeline" not in sys.modules:
        monkeypatch.setitem(sys.modules, "app.pipeline", types.ModuleType("app.pipeline"))

    fake_docling = types.ModuleType("app.pipeline.docling_parser")

    class ParsedDocument:
        def __init__(self, full_text: str = "") -> None:
            self.full_text = full_text

    fake_docling.ParsedDocument = ParsedDocument
    fake_docling.parse_pdf = lambda source, filename=None: ParsedDocument("")  # noqa: E731
    fake_docling.release_converter = lambda: None  # noqa: E731

    fake_extrator = types.ModuleType("extrator_evidencia")
    fake_extrator.build_llm_input_from_merged_text = lambda text: text  # noqa: E731

    monkeypatch.setitem(sys.modules, "app.pipeline.docling_parser", fake_docling)
    monkeypatch.setitem(sys.modules, "extrator_evidencia", fake_extrator)

    mod_path = Path(__file__).resolve().parents[2] / "AutomatizadorDePlanilha" / "pipeline.py"
    spec = importlib.util.spec_from_file_location("automatizador_planilha_pipeline_mod", mod_path)
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def test_append_row_to_xlsx_atualiza_mesmo_n_interno(monkeypatch, tmp_path: Path) -> None:
    mod = _load_pipeline_module(monkeypatch)
    planilha_path = tmp_path / "planilha.xlsx"

    row_inicial = {col: "N/C" for col in mod.PLANILHA_COLUMNS}
    row_inicial["N interno"] = "2026_04_08_1"
    row_inicial["Status"] = "verde"
    row_inicial["URL DRIVE"] = "https://drive.exemplo/primeira"
    mod.append_row_to_xlsx(planilha_path, row_inicial)

    row_atualizada = {col: "N/C" for col in mod.PLANILHA_COLUMNS}
    row_atualizada["N interno"] = "2026_04_08_1"
    row_atualizada["Status"] = "amarelo"
    row_atualizada["URL DRIVE"] = "https://drive.exemplo/segunda"
    mod.append_row_to_xlsx(planilha_path, row_atualizada)

    wb = load_workbook(planilha_path)
    ws = wb.active
    header_map = {
        str(ws.cell(row=1, column=col_idx).value): col_idx
        for col_idx in range(1, ws.max_column + 1)
    }

    assert ws.max_row == 2
    assert ws.cell(row=2, column=header_map["Status"]).value == "amarelo"
    assert ws.cell(row=2, column=header_map["URL DRIVE"]).value == "https://drive.exemplo/segunda"


def test_extract_n_interno_aceita_pdf_direto(monkeypatch, tmp_path: Path) -> None:
    mod = _load_pipeline_module(monkeypatch)
    pdf_path = tmp_path / "2026_04_08_15_-_edital.pdf"
    pdf_path.write_bytes(b"%PDF-1.4")

    assert mod._extract_n_interno_from_folder(pdf_path) == "2026_04_08_15"
